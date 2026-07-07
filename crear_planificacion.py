import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Planificacion"

DIM_COLORS = {
    1: "E8F4FD", 2: "E8F8F5", 3: "FEF9E7",
    4: "FDEDEC", 5: "F4ECF7", 6: "EAF2F8", 7: "E9F7EF",
}

headers = ["Dimension", "Punto", "Nombre del punto",
           "Visualizaciones incluidas", "Fuente de datos",
           "Replicado", "Actualizado", "Responsable", "Notas"]

header_fill = PatternFill("solid", fgColor="2C3E50")
header_font = Font(bold=True, color="FFFFFF", size=10)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

filas = [
    # dim, dimension, punto, nombre_punto, visualizaciones, fuente, replicado, actualizado, responsable, notas
    # DIM 1
    (1,"1. Ser joven","1.1","Caracteristicas demograficas",
     "KPI total jovenes y % poblacion · Jovenes por zona · Piramide edad/sexo · Distribucion por localidad · Proyeccion 2024-2030",
     "DANE - Proyecciones CNPV 2018","Si","Si","","Proyecciones fijas, no requieren actualizacion periodica"),

    # DIM 2
    (2,"2. Educacion","2.1","Educacion media",
     "KPIs cobertura bruta/neta/aprobacion · KPIs reprobacion/repitencia/desercion · Lineas temporales",
     "MEN - datos.gov.co (preescolar, basica y media)","Si","Pendiente","","Actualizar cuando MEN publique datos 2025"),
    (2,"2. Educacion","2.2","Educacion superior",
     "KPIs matriculas Bogota/Colombia/% participacion · Barras por nivel de formacion · Lineas historicas",
     "MEN - matriculas por municipio ES","Si","Pendiente","",""),

    # DIM 3
    (3,"3. Inclusion Productiva","3.1","Panorama laboral juvenil",
     "Arbol PET/ocupados/desocupados · KPIs TGP/TO/TD · Serie historica Bogota · Comparacion 13 ciudades · Tabla por entidad",
     "DANE - GEIH Mercado laboral juvenil","Si","Pendiente","","Actualizar con trimestre mas reciente 2025"),

    # DIM 4
    (4,"4. Salud Integral y Autocuidado","4.1","Afiliacion al SGSSS",
     "KPI total afiliados · Distribucion por regimen (contributivo/subsidiado)",
     "Por definir","No","Pendiente","","Fuente no disponible aun"),
    (4,"4. Salud Integral y Autocuidado","4.2","Discapacidad certificada",
     "KPI total jovenes con discapacidad · Barras por tipo · Barras por localidad",
     "OSB - SaludData","Si","Pendiente","","Categorias no excluyentes"),
    (4,"4. Salud Integral y Autocuidado","4.3","Natalidad en jovenes",
     "KPIs tasa por mil y nacidos vivos · Barras por grupo de edad madre · Linea historica · Barras por localidad",
     "OSB + DANE Proyecciones CNPV 2018","Si","Pendiente","","Grupos: 15-19, 20-24, 25-28"),

    # DIM 5
    (5,"5. Cultura, Recreacion y Deporte","5.1","Satisfaccion con oferta cultural y recreativa",
     "KPI satisfaccion recreativa/deportiva (CUL_1) · KPI satisfaccion cultural (CUL_2) · Libros leidos (CUL_3)",
     "Bogota Como Vamos - EPC 2025, microdatos SPSS","Si","Si","","Datos 2025 disponibles"),
    (5,"5. Cultura, Recreacion y Deporte","5.2","Participacion en actividades",
     "% participacion: Cine, Ferias, Parques, Ciclovia, etc.",
     "Bogota Como Vamos - fuente adicional pendiente","No","Pendiente","","NO disponible en microdatos actuales. Conseguir fuente de BCV."),

    # DIM 6
    (6,"6. Paz, Convivencia y Justicia","6.1","Percepcion de seguridad y convivencia",
     "KPIs victima de delito / percepcion barrio / percepcion ciudad · Barras problemas del barrio · Barras apiladas por variable",
     "Bogota Como Vamos - EPC 2025, microdatos SPSS","No","Si","","Script listo, falta construir index.html"),

    # DIM 7
    (7,"7. Habitat","7.1","Percepcion sobre habitat y medio ambiente",
     "KPIs satisfaccion agua (AMB_2) / ruido (AMB_3) / Bogota como lugar para vivir (GOB_3) · Barras apiladas comparativo",
     "Bogota Como Vamos - EPC 2025, microdatos SPSS","No","Si","","Script listo, falta construir index.html. GOB_3 es proxy de satisfaccion con vivienda."),
]

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, fila in enumerate(filas, 2):
    dim_num = fila[0]
    valores = list(fila[1:])  # sin el numero de dim
    fill = PatternFill("solid", fgColor=DIM_COLORS[dim_num])
    for col_idx, val in enumerate(valores, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = Font(size=9)
    ws.row_dimensions[row_idx].height = 52

ws.column_dimensions["A"].width = 24   # Dimension
ws.column_dimensions["B"].width = 8    # Punto
ws.column_dimensions["C"].width = 26   # Nombre del punto
ws.column_dimensions["D"].width = 52   # Visualizaciones
ws.column_dimensions["E"].width = 30   # Fuente
ws.column_dimensions["F"].width = 12   # Replicado
ws.column_dimensions["G"].width = 12   # Actualizado
ws.column_dimensions["H"].width = 14   # Responsable
ws.column_dimensions["I"].width = 38   # Notas

out = "G:/Mi unidad/CH_projects/SDIS/tablero-cij/planificacion-tablero-cij.xlsx"
wb.save(out)
print("Guardado:", out)
