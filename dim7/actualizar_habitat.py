# actualizar_habitat.py — Dimensión 7: Hábitat
# Procesa dos fuentes distintas y genera el JSON de hábitat en jóvenes:
#
# 1. Microdatos SPSS de la Encuesta de Percepción Ciudadana (Bogotá Cómo Vamos,
#    jóvenes 18-25 años, sin ponderar):
#      AMB_2: Satisfacción calidad del agua (ríos, quebradas, humedales...)
#      AMB_3: Satisfacción ruido de la ciudad
#      GOB_3: Satisfacción con Bogotá como lugar para vivir
#      SER_1: Satisfacción con el servicio de agua potable
#      SER_4: Satisfacción con aseo y recolección de basuras
#      CCC_3: Percepción de respeto por las normas ambientales en la ciudad
#
# 2. Microdatos del módulo "Cultura Ambiental" de la Encuesta Bienal de
#    Culturas 2025 (jóvenes 13-28 años, ponderados por factor de expansión):
#      P78: ¿Separa los residuos que se generan en su casa?
#      P85: ¿Lava o limpia los residuos aprovechables antes de botarlos?
#      P87: ¿Se asegura de que los residuos orgánicos sean aprovechados en compostaje?
#
# Nota: son dos encuestas distintas con rangos de edad distintos (18-25 vs.
# 13-28). El tablero web debe mostrar esto con claridad, no mezclarlas como
# si fueran una sola fuente.
#
# Variable NO disponible en microdatos (pendiente):
#   - Satisfacción con la vivienda que habita (se usa GOB_3 como proxy)
#
# Uso:
#   python dim7/actualizar_habitat.py
#
# Dependencias: pyreadstat, openpyxl (pip install pyreadstat openpyxl)

import os
import json
import pyreadstat
import openpyxl

# ============================================================
# Configuración
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUENTES_DIR = os.path.join(SCRIPT_DIR, 'fuentes', 'Bogota como vamos')
AMBIENTE_XLSX = os.path.join(
    os.path.dirname(SCRIPT_DIR), 'dim5', 'fuentes',
    'Encuesta Bienal de Culturas', 'Ambiente', 'm194_datos_consolidados.xlsx'
)
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

# Filtro de edad (jóvenes) — Bogotá Cómo Vamos
EDAD_MIN = 18
EDAD_MAX = 25

# Filtro de edad (jóvenes) — Encuesta Bienal de Culturas (módulo Ambiental)
EDAD_MIN_AMB = 13
EDAD_MAX_AMB = 28

# Columnas confirmadas contra el codebook de m194_datos_consolidados.xlsx
COLS_AMBIENTAL = {
    'edad': 'SI',    # D3
    'factor': 'TQ',  # FACTOR
    'separa_residuos': 'ON',      # P78
    'limpia_reciclables': 'PG',   # P85
    'compostaje': 'PI',           # P87
    'satisfaccion_vivienda': 'LX',  # P61.3 — "Su vivienda" (escala 1-4)
}


# ============================================================
# Funciones
# ============================================================

def encontrar_spss(carpeta):
    """Busca el archivo SPSS de microdatos."""
    for f in os.listdir(carpeta):
        if f.lower().startswith('microdatos') and not f.endswith('.xlsx'):
            return os.path.join(carpeta, f)
        if f.lower().endswith('.sav'):
            return os.path.join(carpeta, f)
    return None


def calcular_satisfaccion(serie):
    """Calcula porcentajes de satisfacción agrupados en 3 categorías.
    Escala: 1-2 insatisfecho, 3 ni/ni, 4-5 satisfecho. Excluye 90 (No sabe).
    """
    datos = serie.dropna()
    datos = datos[datos <= 5]  # Excluir "No sabe" (90)
    n = len(datos)
    if n == 0:
        return {'satisfecho': 0, 'ni_satisfecho_ni_insatisfecho': 0, 'insatisfecho': 0, 'n': 0}

    satisfecho = (datos >= 4).sum()
    ni = (datos == 3).sum()
    insatisfecho = (datos <= 2).sum()

    return {
        'satisfecho': round(satisfecho / n * 100, 2),
        'ni_satisfecho_ni_insatisfecho': round(ni / n * 100, 2),
        'insatisfecho': round(insatisfecho / n * 100, 2),
        'n': n,
    }


def procesar(ruta_spss):
    """Lee el SPSS, filtra jóvenes y calcula indicadores de hábitat."""
    df, meta = pyreadstat.read_sav(ruta_spss)

    # Filtrar jóvenes 18-25
    jovenes = df[(df['DMO_3_1'] >= EDAD_MIN) & (df['DMO_3_1'] <= EDAD_MAX)].copy()
    n = len(jovenes)
    print(f'  Jóvenes 18-25: {n} de {len(df)} encuestados')

    # --- 1. Satisfacción calidad del agua (AMB_2) ---
    agua = calcular_satisfaccion(jovenes['AMB_2'])

    # --- 2. Satisfacción ruido de la ciudad (AMB_3) ---
    ruido = calcular_satisfaccion(jovenes['AMB_3'])

    # --- 3. Satisfacción con Bogotá como lugar para vivir (GOB_3) ---
    vivienda = calcular_satisfaccion(jovenes['GOB_3'])

    # --- 4. Satisfacción con el servicio de agua potable (SER_1) ---
    agua_servicio = calcular_satisfaccion(jovenes['SER_1'])

    # --- 5. Satisfacción con aseo y recolección de basuras (SER_4) ---
    aseo = calcular_satisfaccion(jovenes['SER_4'])

    # --- 6. Percepción de respeto por las normas ambientales (CCC_3) ---
    normas_ambientales = calcular_satisfaccion(jovenes['CCC_3'])

    return {
        'fuente': 'Encuesta de Percepción Ciudadana - Bogotá Cómo Vamos - 2025',
        'poblacion': f'Jóvenes de {EDAD_MIN} a {EDAD_MAX} años de Bogotá',
        'n_jovenes': n,
        'satisfaccion_agua': agua,
        'satisfaccion_ruido': ruido,
        'satisfaccion_vivienda': vivienda,
        'satisfaccion_agua_servicio': agua_servicio,
        'satisfaccion_aseo': aseo,
        'percepcion_normas_ambientales': normas_ambientales,
        'notas': 'satisfaccion_vivienda aquí es GOB_3 (satisfacción con Bogotá como lugar para vivir en general, no con la vivienda específica). La satisfacción con la vivienda propiamente dicha está en acciones_ambientales.satisfaccion_vivienda (otra encuesta, ver sección 7.2).',
    }


def procesar_acciones_ambientales(ruta_xlsx):
    """Lee el módulo Cultura Ambiental de la Bienal de Culturas, filtra
    jóvenes 13-28 y calcula % ponderado de Sí para cada acción concreta."""
    wb = openpyxl.load_workbook(ruta_xlsx, data_only=True)
    ws = wb['datos']
    col_letters = [c.column_letter for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {letra: i for i, letra in enumerate(col_letters)}
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    claves = ['separa_residuos', 'limpia_reciclables', 'compostaje']
    peso_si = {clave: 0.0 for clave in claves}
    peso_vivienda = {1: 0.0, 2: 0.0, 3: 0.0, 4: 0.0}
    total = 0.0
    total_vivienda = 0.0
    n = 0

    for fila in filas:
        edad = fila[idx[COLS_AMBIENTAL['edad']]]
        if edad is None:
            continue
        edad = int(edad)
        if edad < EDAD_MIN_AMB or edad > EDAD_MAX_AMB:
            continue
        factor = fila[idx[COLS_AMBIENTAL['factor']]] or 0
        total += factor
        n += 1
        for clave in claves:
            valor = fila[idx[COLS_AMBIENTAL[clave]]]
            if valor == 1:  # 1 = Sí, 2 = No (confirmado en etiquetas_respuestas)
                peso_si[clave] += factor

        valor_vivienda = fila[idx[COLS_AMBIENTAL['satisfaccion_vivienda']]]
        if valor_vivienda in (1, 2, 3, 4):
            peso_vivienda[valor_vivienda] += factor
            total_vivienda += factor

    print(f'  Jóvenes 13-28 (módulo Ambiental): {n} encuestados, {total:,.0f} personas ponderadas'.replace(',', '.'))

    satisfaccion_vivienda = {
        'nada': round(peso_vivienda[1] / total_vivienda * 100, 1) if total_vivienda else 0,
        'poco': round(peso_vivienda[2] / total_vivienda * 100, 1) if total_vivienda else 0,
        'satisfecho': round(peso_vivienda[3] / total_vivienda * 100, 1) if total_vivienda else 0,
        'muy_satisfecho': round(peso_vivienda[4] / total_vivienda * 100, 1) if total_vivienda else 0,
    }

    return {
        'fuente': 'Encuesta Bienal de Culturas 2025 — módulo Cultura Ambiental',
        'poblacion': f'Jóvenes de {EDAD_MIN_AMB} a {EDAD_MAX_AMB} años de Bogotá',
        'n_jovenes': n,
        'separa_residuos': round(peso_si['separa_residuos'] / total * 100, 1) if total else 0,
        'limpia_reciclables': round(peso_si['limpia_reciclables'] / total * 100, 1) if total else 0,
        'compostaje': round(peso_si['compostaje'] / total * 100, 1) if total else 0,
        'satisfaccion_vivienda': satisfaccion_vivienda,
        'nota': 'Esta sección viene de una encuesta distinta (Bienal de Culturas, módulo Cultura Ambiental), '
                'con un rango de edad diferente (13-28 años) al resto de la página (18-25 años, Bogotá Cómo Vamos). '
                'La satisfacción con la vivienda reemplaza el proxy anterior (GOB_3), que medía satisfacción con '
                'Bogotá en general, no con la vivienda específicamente.',
    }


def guardar_json(datos):
    os.makedirs(DATA_DIR, exist_ok=True)
    ruta = os.path.join(DATA_DIR, 'habitat.json')
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)
    print(f'  habitat.json generado')
    return ruta


def main():
    print('=' * 60)
    print('Actualización — Dimensión 7: Hábitat')
    print('=' * 60)

    archivo = encontrar_spss(FUENTES_DIR)
    if not archivo:
        print('ERROR: No se encontró archivo SPSS de microdatos')
        return
    print(f'  Archivo: {os.path.basename(archivo)}')

    print('\nProcesando Bogotá Cómo Vamos...')
    datos = procesar(archivo)

    print('\nProcesando módulo Cultura Ambiental (Bienal de Culturas)...')
    if not os.path.exists(AMBIENTE_XLSX):
        print(f'  ERROR: No se encontró {AMBIENTE_XLSX}')
        return
    datos['acciones_ambientales'] = procesar_acciones_ambientales(AMBIENTE_XLSX)

    print('\nGenerando JSON...')
    ruta = guardar_json(datos)

    # Verificación
    print(f'\n{"=" * 60}')
    print('VERIFICACIÓN:')
    print(f'{"=" * 60}')
    print(f'  Jóvenes encuestados (Bogotá Cómo Vamos): {datos["n_jovenes"]}')
    claves = [
        ('satisfaccion_agua', 'Calidad del agua'),
        ('satisfaccion_ruido', 'Ruido'),
        ('satisfaccion_vivienda', 'Bogotá como lugar para vivir'),
        ('satisfaccion_agua_servicio', 'Servicio de agua potable'),
        ('satisfaccion_aseo', 'Aseo y recolección de basuras'),
        ('percepcion_normas_ambientales', 'Percepción de respeto a normas ambientales'),
    ]
    for key, label in claves:
        d = datos[key]
        print(f'\n  {label}:')
        print(f'    Satisfecho {d["satisfecho"]}% / Ni {d["ni_satisfecho_ni_insatisfecho"]}% / Insatisfecho {d["insatisfecho"]}%  (n={d["n"]})')

    acc = datos['acciones_ambientales']
    print(f'\n  Acciones ambientales (jóvenes 13-28, Bienal de Culturas):')
    print(f'    Separa residuos: {acc["separa_residuos"]}%')
    print(f'    Limpia reciclables: {acc["limpia_reciclables"]}%')
    print(f'    Compostaje: {acc["compostaje"]}%')
    sv = acc['satisfaccion_vivienda']
    print(f'    Satisfacción con la vivienda: Nada {sv["nada"]}% / Poco {sv["poco"]}% / Satisfecho {sv["satisfecho"]}% / Muy satisfecho {sv["muy_satisfecho"]}%')

    print(f'\nArchivo: {ruta}')


if __name__ == '__main__':
    main()
