# actualizar.py — Dimensión 1: Ser joven
# Script para procesar las proyecciones de población de Bogotá y generar
# los JSONs que usa la página web.
#
# Desde 2026-07-15 la fuente es la SDP (convenio SDP-DANE, actualización
# post-COVID del CNPV 2018), no la página del DANE. Este script ya no lee
# el archivo de la SDP directamente: lee la fuente estándar
# fuentes/fuente_dim1_poblacion.xlsx, que genera dim1/generar_fuente.py.
#
# Uso:
#   1. Si hay archivo nuevo de la SDP, correr primero:
#      python dim1/generar_fuente.py
#      (ver instrucciones de descarga en ese script o en INSTRUCCIONES.md)
#
#   2. Correr este script:
#      python dim1/actualizar.py
#
#   3. Los archivos JSON en dim1/data/ se actualizan automáticamente.

import os
import re
import json
import openpyxl

# ============================================================
# Configuración
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUENTES_DIR = os.path.join(SCRIPT_DIR, 'fuentes')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

# Fuente estándar de la dimensión (nombre fijo, la genera generar_fuente.py)
FUENTE = os.path.join(FUENTES_DIR, 'fuente_dim1_poblacion.xlsx')

# Rango de edad para "joven" según Ley 1622 (Estatuto de Ciudadanía Juvenil)
EDAD_MIN = 14
EDAD_MAX = 28

# Años a incluir en el JSON (no necesitamos todo 2018-2035 para el tablero)
ANIO_MIN = 2018
ANIO_MAX = 2035

# Edades de la pirámide poblacional general: 0 a 100 por edad simple
# (100 agrupa "100 y más"). Se usa edad simple y no quinquenios para poder
# resaltar exacto el rango de juventud 14-28 (los quinquenios lo parten).
EDADES_PIRAMIDE = [str(e) for e in range(101)]

# Nombre de la hoja de datos en la fuente estándar
HOJA_PRINCIPAL = 'Localidades'

# Fila de los headers: en la fuente estándar siempre es la 1
FILA_HEADERS = 1


# ============================================================
# Funciones
# ============================================================

def leer_headers(ws, fila):
    """Lee los nombres de columna de la fila de headers."""
    headers = []
    for cell in ws[fila]:
        headers.append(cell.value)
    return headers


def extraer_datos(ws, headers):
    """Extrae todos los datos de la hoja en una lista de diccionarios."""
    # Índices de columnas clave
    idx_loc = headers.index('COD_LOC')
    idx_nom = headers.index('NOM_LOC')
    idx_area = headers.index('AREA') if 'AREA' in headers else headers.index('ÁREA')
    # El año puede venir como AÑO o A?O por encoding
    idx_anio = None
    for i, h in enumerate(headers):
        if h and ('AÑO' in str(h).upper() or 'A\ufffdO' in str(h) or 'ANO' in str(h).upper()
                  or str(h).upper().startswith('A') and str(h).upper().endswith('O') and len(str(h)) <= 4):
            idx_anio = i
            break
    if idx_anio is None:
        idx_anio = 3  # posición por defecto

    # Mapear columnas de edad por sexo
    # Formato: Hombres_14, Mujeres_14, Total_14
    col_hombres = {}  # edad -> índice de columna
    col_mujeres = {}
    col_total = {}

    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h)
        for prefix, destino in [('Hombres_', col_hombres), ('Mujeres_', col_mujeres), ('Total_', col_total)]:
            if h.startswith(prefix):
                edad_str = h[len(prefix):]
                if edad_str.isdigit():
                    destino[int(edad_str)] = i
                elif edad_str == '100 y más':
                    # se guarda como edad 100 para la pirámide general
                    destino[100] = i

    # Índices de totales generales
    idx_total_h = None
    idx_total_m = None
    idx_total = None
    for i, h in enumerate(headers):
        if h == 'TOTAL HOMBRES':
            idx_total_h = i
        elif h == 'TOTAL MUJERES':
            idx_total_m = i
        elif h == 'TOTAL':
            idx_total = i

    registros = []
    for row in ws.iter_rows(min_row=FILA_HEADERS + 1, values_only=True):
        if not row[idx_loc]:
            continue

        anio = int(row[idx_anio]) if row[idx_anio] else None
        if not anio or anio < ANIO_MIN or anio > ANIO_MAX:
            continue

        # Población joven por edad y sexo
        jovenes_h = {}
        jovenes_m = {}
        jovenes_t = {}
        for edad in range(EDAD_MIN, EDAD_MAX + 1):
            if edad in col_hombres:
                jovenes_h[edad] = int(row[col_hombres[edad]] or 0)
            if edad in col_mujeres:
                jovenes_m[edad] = int(row[col_mujeres[edad]] or 0)
            if edad in col_total:
                jovenes_t[edad] = int(row[col_total[edad]] or 0)

        # Pirámide general: toda la población por edad simple y sexo
        piramide_h = {g: 0 for g in EDADES_PIRAMIDE}
        piramide_m = {g: 0 for g in EDADES_PIRAMIDE}
        for edad, col in col_hombres.items():
            piramide_h[str(min(edad, 100))] += int(row[col] or 0)
        for edad, col in col_mujeres.items():
            piramide_m[str(min(edad, 100))] += int(row[col] or 0)

        total_jovenes_h = sum(jovenes_h.values())
        total_jovenes_m = sum(jovenes_m.values())
        total_jovenes = sum(jovenes_t.values())
        total_poblacion = int(row[idx_total] or 0) if idx_total else 0
        total_pob_h = int(row[idx_total_h] or 0) if idx_total_h else 0
        total_pob_m = int(row[idx_total_m] or 0) if idx_total_m else 0

        registros.append({
            'cod_loc': str(row[idx_loc]).zfill(2),
            'localidad': row[idx_nom],
            'area': row[idx_area],
            'anio': anio,
            'jovenes_hombres': total_jovenes_h,
            'jovenes_mujeres': total_jovenes_m,
            'jovenes_total': total_jovenes,
            'poblacion_total': total_poblacion,
            'poblacion_hombres': total_pob_h,
            'poblacion_mujeres': total_pob_m,
            'por_edad_hombres': jovenes_h,
            'por_edad_mujeres': jovenes_m,
            'piramide_hombres': piramide_h,
            'piramide_mujeres': piramide_m,
        })

    return registros


def generar_jsons(registros):
    """Genera los archivos JSON para el tablero web."""
    os.makedirs(DATA_DIR, exist_ok=True)

    # 1. Resumen por año para Bogotá (suma de todas las localidades, area=Total)
    resumen = {}
    for r in registros:
        if r['area'] != 'Total':
            continue
        anio = r['anio']
        if anio not in resumen:
            resumen[anio] = {
                'anio': anio,
                'jovenes_hombres': 0, 'jovenes_mujeres': 0, 'jovenes_total': 0,
                'poblacion_total': 0,
                'por_edad_hombres': {}, 'por_edad_mujeres': {},
                'piramide_hombres': {g: 0 for g in EDADES_PIRAMIDE},
                'piramide_mujeres': {g: 0 for g in EDADES_PIRAMIDE},
                'zona_cabecera': 0, 'zona_rural': 0,
            }
        resumen[anio]['jovenes_hombres'] += r['jovenes_hombres']
        resumen[anio]['jovenes_mujeres'] += r['jovenes_mujeres']
        resumen[anio]['jovenes_total'] += r['jovenes_total']
        resumen[anio]['poblacion_total'] += r['poblacion_total']
        # Acumular la pirámide general por edad simple
        for g in EDADES_PIRAMIDE:
            resumen[anio]['piramide_hombres'][g] += r['piramide_hombres'][g]
            resumen[anio]['piramide_mujeres'][g] += r['piramide_mujeres'][g]
        # Acumular por edad
        for edad in range(EDAD_MIN, EDAD_MAX + 1):
            e_str = str(edad)
            resumen[anio]['por_edad_hombres'][e_str] = (
                resumen[anio]['por_edad_hombres'].get(e_str, 0) + r['por_edad_hombres'].get(edad, 0)
            )
            resumen[anio]['por_edad_mujeres'][e_str] = (
                resumen[anio]['por_edad_mujeres'].get(e_str, 0) + r['por_edad_mujeres'].get(edad, 0)
            )

    # Agregar zona (cabecera vs rural) desde los registros por área
    for r in registros:
        anio = r['anio']
        if anio not in resumen:
            continue
        if r['area'] == 'Cabecera Municipal':
            resumen[anio]['zona_cabecera'] += r['jovenes_total']
        elif r['area'] == 'Centro Poblado y Rural Disperso':
            # etiqueta de la SDP (el archivo viejo del DANE decía "Centros Poblados...")
            resumen[anio]['zona_rural'] += r['jovenes_total']

    resumen_list = sorted(resumen.values(), key=lambda x: x['anio'])

    ruta_resumen = os.path.join(DATA_DIR, 'resumen_bogota.json')
    with open(ruta_resumen, 'w', encoding='utf-8') as f:
        json.dump(resumen_list, f, ensure_ascii=False, indent=2)
    print(f'  resumen_bogota.json: {len(resumen_list)} años')

    # 2. Datos por localidad (area=Total), con edades 14-28 por sexo y zona
    # para que el zoom de juventud y la gráfica de zona respondan al filtro
    zonas = {}
    for r in registros:
        clave = (r['cod_loc'], r['anio'])
        if r['area'] == 'Cabecera Municipal':
            zonas.setdefault(clave, [0, 0])[0] += r['jovenes_total']
        elif r['area'] == 'Centro Poblado y Rural Disperso':
            zonas.setdefault(clave, [0, 0])[1] += r['jovenes_total']

    localidades = []
    for r in registros:
        if r['area'] != 'Total':
            continue
        zona = zonas.get((r['cod_loc'], r['anio']), [0, 0])
        localidades.append({
            'cod_loc': r['cod_loc'],
            'localidad': r['localidad'],
            'anio': r['anio'],
            'jovenes_hombres': r['jovenes_hombres'],
            'jovenes_mujeres': r['jovenes_mujeres'],
            'jovenes_total': r['jovenes_total'],
            'poblacion_total': r['poblacion_total'],
            'por_edad_hombres': r['por_edad_hombres'],
            'por_edad_mujeres': r['por_edad_mujeres'],
            'zona_cabecera': zona[0],
            'zona_rural': zona[1],
        })

    ruta_loc = os.path.join(DATA_DIR, 'localidades.json')
    # compacto (sin indentación): el archivo creció con las edades por localidad
    with open(ruta_loc, 'w', encoding='utf-8') as f:
        json.dump(localidades, f, ensure_ascii=False, separators=(',', ':'))
    locs_unicas = len(set(r['localidad'] for r in localidades))
    print(f'  localidades.json: {len(localidades)} registros ({locs_unicas} localidades)')

    return ruta_resumen, ruta_loc


def actualizar_fallback_html(ruta_resumen, ruta_loc):
    """Regenera los datos embebidos de respaldo en index.html.
    El tablero publicado lee los JSON de data/, pero al abrir el HTML local
    sin servidor usa los datos embebidos (RESUMEN_FALLBACK y
    LOCALIDADES_FALLBACK). Se reescriben aquí en cada actualización para que
    nunca queden con cifras viejas frente a los JSON."""
    ruta_html = os.path.join(SCRIPT_DIR, 'index.html')
    with open(ruta_html, 'r', encoding='utf-8') as f:
        html = f.read()

    bloques = [('RESUMEN_FALLBACK', ruta_resumen), ('LOCALIDADES_FALLBACK', ruta_loc)]
    for nombre, ruta_json in bloques:
        with open(ruta_json, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        # compacto: los datos embebidos no necesitan ser legibles y así pesa menos
        nuevo = f'const {nombre} = ' + json.dumps(datos, ensure_ascii=False, separators=(',', ':')) + ';'
        patron = re.compile(r'const ' + nombre + r' = \[.*?\];', re.DOTALL)
        if not patron.search(html):
            print(f'  AVISO: no se encontró el bloque {nombre} en index.html — no se actualizó.')
            continue
        html = patron.sub(lambda m: nuevo, html, count=1)
        print(f'  {nombre} actualizado en index.html')

    with open(ruta_html, 'w', encoding='utf-8') as f:
        f.write(html)


def actualizar_fallback_index_raiz(ruta_resumen):
    """Regenera los datos embebidos del index.html de la RAÍZ del tablero.
    El index raíz muestra los KPI de jóvenes leyendo dim1/data/resumen_bogota.json,
    pero al abrirlo local sin servidor usa su propio RESUMEN_FALLBACK embebido
    (solo año, jóvenes y población). Se reescribe aquí para que el index y la
    dimensión 1 nunca muestren cifras distintas."""
    ruta_html = os.path.join(os.path.dirname(SCRIPT_DIR), 'index.html')
    if not os.path.exists(ruta_html):
        print('  AVISO: no se encontró el index.html raíz — no se actualizó.')
        return
    with open(ruta_resumen, 'r', encoding='utf-8') as f:
        resumen = json.load(f)
    reducido = [{'anio': r['anio'], 'jovenes_total': r['jovenes_total'],
                 'poblacion_total': r['poblacion_total']} for r in resumen]

    with open(ruta_html, 'r', encoding='utf-8') as f:
        html = f.read()
    nuevo = 'const RESUMEN_FALLBACK = ' + json.dumps(reducido, ensure_ascii=False, separators=(',', ':')) + ';'
    patron = re.compile(r'const RESUMEN_FALLBACK = \[.*?\];', re.DOTALL)
    if not patron.search(html):
        print('  AVISO: no se encontró RESUMEN_FALLBACK en el index raíz — no se actualizó.')
        return
    html = patron.sub(lambda m: nuevo, html, count=1)
    with open(ruta_html, 'w', encoding='utf-8') as f:
        f.write(html)
    print('  RESUMEN_FALLBACK actualizado en el index raíz')


def main():
    print('=' * 60)
    print('Actualización de datos — Dimensión 1: Ser joven')
    print('=' * 60)

    if not os.path.exists(FUENTE):
        print(f'ERROR: no existe {FUENTE}')
        print('Correr primero: python dim1/generar_fuente.py')
        return
    print(f'\nFuente estándar: {os.path.basename(FUENTE)}')

    print('\nLeyendo datos...')
    wb = openpyxl.load_workbook(FUENTE, read_only=True, data_only=True)
    ws = wb[HOJA_PRINCIPAL]

    headers = leer_headers(ws, FILA_HEADERS)
    registros = extraer_datos(ws, headers)
    wb.close()
    print(f'  {len(registros)} registros extraídos')

    print('\nGenerando archivos JSON...')
    ruta_resumen, ruta_loc = generar_jsons(registros)

    print('\nActualizando datos embebidos en index.html...')
    actualizar_fallback_html(ruta_resumen, ruta_loc)
    actualizar_fallback_index_raiz(ruta_resumen)

    # Verificación rápida
    with open(ruta_resumen, 'r', encoding='utf-8') as f:
        resumen = json.load(f)
    ultimo = [r for r in resumen if r['anio'] == 2024]
    if ultimo:
        d = ultimo[0]
        print(f'\n{"=" * 60}')
        print(f'VERIFICACIÓN — Año 2024:')
        print(f'{"=" * 60}')
        print(f'  Jóvenes (14-28): {d["jovenes_total"]:,.0f}'.replace(',', '.'))
        pct = d['jovenes_total'] / d['poblacion_total'] * 100 if d['poblacion_total'] else 0
        print(f'  % de la población: {pct:.2f}%'.replace('.', ','))
        print(f'  Zona cabecera: {d["zona_cabecera"]:,.0f}'.replace(',', '.'))
        print(f'  Zona rural: {d["zona_rural"]:,.0f}'.replace(',', '.'))
        print(f'  Hombres: {d["jovenes_hombres"]:,.0f}'.replace(',', '.'))
        print(f'  Mujeres: {d["jovenes_mujeres"]:,.0f}'.replace(',', '.'))

    print(f'\nArchivos actualizados:')
    print(f'  - {ruta_resumen}')
    print(f'  - {ruta_loc}')
    print(f'\nEl tablero web se actualizará automáticamente al hacer push.')


if __name__ == '__main__':
    main()
