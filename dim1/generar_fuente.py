# generar_fuente.py — Dimensión 1: Ser joven
# Convierte el archivo de proyecciones de población por localidad que publica
# la SDP (convenio SDP-DANE, actualización post-COVID del CNPV 2018) en la
# fuente estándar del tablero: fuentes/fuente_dim1_poblacion.xlsx
#
# Por qué existe este script: el DANE es la fuente oficial de las proyecciones
# pero no ha actualizado su página; la SDP sí publica las actualizaciones del
# convenio SDP-DANE. Cada vez que la SDP actualiza cambia el nombre del archivo,
# así que este script toma el archivo más reciente de fuentes/SDP-Dane/ y genera
# siempre el mismo archivo estándar, con nombre y formato fijos. El resto del
# tablero (actualizar.py) solo conoce la fuente estándar.
#
# Uso:
#   1. Descargar de la página de la SDP (menú "Cifras población", actualización
#      más reciente) el archivo de población a nivel de Localidad:
#      https://www.sdp.gov.co/gestion-estudios-estrategicos/informacion-estadisticas/censo-2018-act-agosto2025/proyecciones-de-poblacion
#
#   2. Guardarlo en dim1/fuentes/SDP-Dane/ (sin borrar el anterior).
#
#   3. Correr este script desde la raíz del repo:
#      python dim1/generar_fuente.py
#
#   4. Se regenera dim1/fuentes/fuente_dim1_poblacion.xlsx
#      Después correr: python dim1/actualizar.py

import os
import re
import sys
import unicodedata
from datetime import date

import openpyxl
from openpyxl.styles import Font

# ============================================================
# Configuración
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SDP_DIR = os.path.join(SCRIPT_DIR, 'fuentes', 'SDP-Dane')
FUENTE_SALIDA = os.path.join(SCRIPT_DIR, 'fuentes', 'fuente_dim1_poblacion.xlsx')

# Página del corte vigente (actualización agosto/2025, contrato 500 de 2025
# DANE-SDP-RMBC). Ojo: la SDP crea una página nueva por cada corte — si este
# enlace queda viejo, entrar por el menú "Cifras población" del sitio de la SDP.
URL_SDP = ('https://www.sdp.gov.co/gestion-estudios-estrategicos/'
           'informacion-estadisticas/censo-2018-act-agosto2025/proyecciones-de-poblacion')

# Áreas que trae el archivo de la SDP. Si un archivo nuevo trae otras etiquetas,
# el script se detiene para que se revise (evita sumar mal los totales).
AREAS_ESPERADAS = {'Cabecera Municipal', 'Centro Poblado y Rural Disperso'}

# Edades de las columnas: 0 a 99 y "100 y más"
EDADES = [str(e) for e in range(100)] + ['100 y más']
PREFIJOS = ['Hombres_', 'Mujeres_', 'Total_']


# ============================================================
# Funciones
# ============================================================

def normalizar(texto):
    """Pasa un texto a minúsculas, sin tildes y con espacios simples,
    para comparar sin sorpresas (la SDP a veces deja espacios dobles)."""
    if texto is None:
        return ''
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', texto)


def parsear_columna_edad(header):
    """Reconoce una columna de edad en cualquiera de las variantes que ha
    usado la SDP y devuelve (prefijo estándar, edad) o None si no es de edad.
      - corte 2025-03: 'Hombres_14', 'Total_100 y más'
      - corte 2025-12: 'Hombres_14', 'Total 14 años', 'Total 100 años  y más'
    """
    series = {'hombres': 'Hombres_', 'mujeres': 'Mujeres_', 'total': 'Total_'}
    n = normalizar(header)
    m = re.match(r'^(hombres|mujeres|total)[ _](\d{1,2})( anos?)?$', n)
    if m:
        return series[m.group(1)], str(int(m.group(2)))
    m = re.match(r'^(hombres|mujeres|total)[ _]100( anos?)? y mas$', n)
    if m:
        return series[m.group(1)], '100 y más'
    return None


def encontrar_archivo_sdp(carpeta):
    """Busca en fuentes/SDP-Dane/ el archivo de población por localidad más
    reciente. La SDP cambia el nombre en cada actualización (el prefijo AAAAMM),
    por eso se busca por palabras clave y se toma el más nuevo."""
    candidatos = []
    for f in os.listdir(carpeta):
        nombre = normalizar(f)
        if not f.endswith('.xlsx') or f.startswith('~$'):
            continue
        if 'localidad' in nombre and 'poblacion' in nombre \
                and 'hogares' not in nombre and 'viviendas' not in nombre:
            candidatos.append(f)

    if not candidatos:
        print(f'ERROR: no se encontró ningún archivo de población por localidad en {carpeta}')
        print('Se espera un .xlsx cuyo nombre contenga "localidad" y "poblacion".')
        print(f'Descargarlo de: {URL_SDP}')
        return None

    # Ordena por el prefijo de fecha AAAAMM si existe; si no, por fecha de modificación
    def clave(f):
        m = re.match(r'^(\d{6})', f)
        if m:
            return (1, int(m.group(1)))
        return (0, os.path.getmtime(os.path.join(carpeta, f)))

    candidatos.sort(key=clave)
    elegido = candidatos[-1]
    if len(candidatos) > 1:
        print(f'  Archivos candidatos: {candidatos}')
    return os.path.join(carpeta, elegido)


def encontrar_fila_headers(ws, max_filas=30):
    """Busca la fila de encabezados: la que contiene 'Nombre Localidad' y 'AÑO'.
    La SDP puede mover los encabezados de fila entre versiones."""
    for i, row in enumerate(ws.iter_rows(max_row=max_filas, values_only=True), start=1):
        valores = [normalizar(v) for v in row]
        if 'nombre localidad' in valores and ('año' in valores or 'ano' in valores):
            return i
    return None


def indice_columna(headers_norm, buscados):
    """Devuelve el índice de una columna probando uno o varios nombres
    posibles (la SDP los cambia entre cortes), sin tildes ni mayúsculas."""
    if isinstance(buscados, str):
        buscados = [buscados]
    for buscado in buscados:
        buscado = normalizar(buscado)
        for i, h in enumerate(headers_norm):
            if h == buscado:
                return i
    raise ValueError(f'No se encontró ninguna columna {buscados} en el archivo de la SDP. '
                     'El formato cambió: revisar el archivo y ajustar generar_fuente.py.')


def leer_archivo_sdp(ruta):
    """Lee el archivo de la SDP y devuelve los registros con todas las edades."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # la SDP publica todo en una sola hoja

    fila_headers = encontrar_fila_headers(ws)
    if fila_headers is None:
        wb.close()
        raise ValueError('No se encontró la fila de encabezados (se busca "Nombre Localidad" y "AÑO"). '
                         'El formato del archivo de la SDP cambió: revisarlo y ajustar el script.')
    print(f'  Encabezados en la fila {fila_headers}')

    headers = [str(c.value) if c.value is not None else '' for c in ws[fila_headers]]
    headers_norm = [normalizar(h) for h in headers]

    # Nombres alternativos: la SDP los cambia entre cortes
    i_cod = indice_columna(headers_norm, 'Código Localidad')
    i_nom = indice_columna(headers_norm, 'Nombre Localidad')
    i_area = indice_columna(headers_norm, ['Área', 'Área Geográfica'])
    i_anio = indice_columna(headers_norm, 'AÑO')
    i_tot_h = indice_columna(headers_norm, 'Total Hombres')
    i_tot_m = indice_columna(headers_norm, 'Total Mujeres')
    i_tot = indice_columna(headers_norm, 'Total')

    # Índices de las columnas de edad, en cualquiera de sus variantes de nombre
    col_edad = {}  # (prefijo, edad) -> índice
    for i, h in enumerate(headers):
        clave = parsear_columna_edad(h)
        if clave:
            col_edad[clave] = i

    faltantes = [(p, e) for p in PREFIJOS for e in EDADES if (p, e) not in col_edad]
    if faltantes:
        wb.close()
        raise ValueError(f'Faltan {len(faltantes)} columnas de edad, por ejemplo {faltantes[:5]}. '
                         'El formato del archivo de la SDP cambió: revisarlo y ajustar el script.')

    registros = []
    for row in ws.iter_rows(min_row=fila_headers + 1, values_only=True):
        if row[i_cod] is None:
            continue
        registros.append({
            'cod_loc': str(row[i_cod]).zfill(2),
            'localidad': row[i_nom],
            'area': row[i_area],
            'anio': int(row[i_anio]),
            'edades': {(p, e): row[col_edad[(p, e)]] or 0 for p in PREFIJOS for e in EDADES},
            'total_hombres': row[i_tot_h] or 0,
            'total_mujeres': row[i_tot_m] or 0,
            'total': row[i_tot] or 0,
        })
    wb.close()
    return registros


def validar(registros):
    """Chequeos de consistencia. Si algo no cuadra, el script se detiene:
    las anomalías se revisan, no se corrigen en silencio."""
    errores = []

    areas = {r['area'] for r in registros}
    if not areas <= AREAS_ESPERADAS:
        errores.append(f'Áreas inesperadas: {areas - AREAS_ESPERADAS}. Se esperaban {AREAS_ESPERADAS}.')

    localidades = {(r['cod_loc'], r['localidad']) for r in registros}
    if len(localidades) != 20:
        errores.append(f'Se esperaban 20 localidades y hay {len(localidades)}.')

    anios = sorted({r['anio'] for r in registros})
    if anios != list(range(anios[0], anios[-1] + 1)):
        errores.append(f'La serie de años tiene huecos: {anios}')

    # El TOTAL de cada fila debe ser la suma de las columnas Total_ por edad
    peor = 0
    for r in registros:
        suma = sum(r['edades'][('Total_', e)] for e in EDADES)
        peor = max(peor, abs(suma - r['total']))
    if peor > 1:
        errores.append(f'El TOTAL por fila no cuadra con la suma por edades (diferencia máxima: {peor}).')

    if errores:
        print('\nERRORES de consistencia en el archivo de la SDP:')
        for e in errores:
            print(f'  - {e}')
        print('Revisar el archivo con Carolina antes de continuar.')
        sys.exit(1)

    print(f'  Consistencia OK: {len(localidades)} localidades, años {anios[0]}-{anios[-1]}, '
          f'áreas {sorted(areas)}')
    return anios


def agregar_totales(registros):
    """Calcula las filas AREA='Total' (cabecera + rural) por localidad y año.
    El archivo viejo del DANE las traía; el de la SDP no, así que se calculan aquí."""
    grupos = {}
    for r in registros:
        grupos.setdefault((r['cod_loc'], r['localidad'], r['anio']), []).append(r)

    totales = []
    for (cod, nom, anio), filas in grupos.items():
        totales.append({
            'cod_loc': cod,
            'localidad': nom,
            'area': 'Total',
            'anio': anio,
            'edades': {k: sum(f['edades'][k] for f in filas) for k in filas[0]['edades']},
            'total_hombres': sum(f['total_hombres'] for f in filas),
            'total_mujeres': sum(f['total_mujeres'] for f in filas),
            'total': sum(f['total'] for f in filas),
        })
    return registros + totales


def escribir_fuente(registros, archivo_origen, ruta_salida):
    """Escribe el Excel estándar: hoja Ficha (documentación) + hoja Localidades (datos)."""
    wb = openpyxl.Workbook()

    # --- Hoja Ficha: para que quien abra el archivo sepa qué es y cómo actualizarlo ---
    ficha = wb.active
    ficha.title = 'Ficha'
    filas_ficha = [
        ('Fuente estándar — Dimensión 1: Ser joven (tablero CIJ)', ''),
        ('', ''),
        ('Contenido', 'Proyecciones y retroproyecciones de población por localidad, '
                      'área y edad simple por sexo. Bogotá D.C.'),
        ('Fuente original', 'DANE - SDP, proyecciones y retroproyecciones de población con '
                            'base censo 2018, contrato interadministrativo DANE-SDP-RMBC'),
        ('Página de descarga', URL_SDP + ' (la SDP crea una página nueva por corte: '
                               'si el enlace quedó viejo, buscar "Cifras población" en sdp.gov.co)'),
        ('Archivo descargado en la SDP', 'Población a nivel de Localidad entre 2005 y 2035'),
        ('Archivo del que se generó', archivo_origen),
        ('Fecha de generación', date.today().isoformat()),
        ('Generado con', 'dim1/generar_fuente.py'),
        ('', ''),
        ('Nota', 'Las filas con AREA = "Total" no vienen en el archivo de la SDP: '
                 'las calcula el script como Cabecera Municipal + Centro Poblado y Rural Disperso.'),
        ('', ''),
        ('Cómo actualizar', '1. Descargar el archivo de localidad de la página de la SDP '
                            'y guardarlo en dim1/fuentes/SDP-Dane/ (el nombre nuevo no importa).'),
        ('', '2. Correr: python dim1/generar_fuente.py'),
        ('', '3. Correr: python dim1/actualizar.py'),
    ]
    for campo, valor in filas_ficha:
        ficha.append([campo, valor])
    ficha['A1'].font = Font(bold=True, size=12)
    for fila in ficha.iter_rows(min_row=3, max_col=1):
        fila[0].font = Font(bold=True)
    ficha.column_dimensions['A'].width = 30
    ficha.column_dimensions['B'].width = 110

    # --- Hoja Localidades: los datos, con encabezados en la fila 1 ---
    ws = wb.create_sheet('Localidades')
    headers = ['COD_LOC', 'NOM_LOC', 'AREA', 'AÑO']
    for p in PREFIJOS:
        headers += [f'{p}{e}' for e in EDADES]
    headers += ['TOTAL HOMBRES', 'TOTAL MUJERES', 'TOTAL']
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = 'E2'

    orden_area = {'Cabecera Municipal': 0, 'Centro Poblado y Rural Disperso': 1, 'Total': 2}
    registros.sort(key=lambda r: (r['cod_loc'], r['anio'], orden_area[r['area']]))
    for r in registros:
        fila = [r['cod_loc'], r['localidad'], r['area'], r['anio']]
        for p in PREFIJOS:
            fila += [r['edades'][(p, e)] for e in EDADES]
        fila += [r['total_hombres'], r['total_mujeres'], r['total']]
        ws.append(fila)

    wb.save(ruta_salida)
    return len(registros)


def main():
    print('=' * 60)
    print('Generación de la fuente estándar — Dimensión 1: Ser joven')
    print('=' * 60)

    print(f'\nBuscando el archivo de la SDP en: {SDP_DIR}')
    archivo = encontrar_archivo_sdp(SDP_DIR)
    if not archivo:
        sys.exit(1)
    print(f'  Archivo: {os.path.basename(archivo)}')

    print('\nLeyendo datos...')
    registros = leer_archivo_sdp(archivo)
    print(f'  {len(registros)} filas leídas')

    print('\nValidando consistencia...')
    validar(registros)

    print('\nCalculando filas de Total (cabecera + rural)...')
    registros = agregar_totales(registros)

    print(f'\nEscribiendo {os.path.basename(FUENTE_SALIDA)}...')
    n = escribir_fuente(registros, os.path.basename(archivo), FUENTE_SALIDA)
    print(f'  {n} filas escritas')

    # Verificación rápida: Bogotá 2024, solo filas Total
    jov = sum(sum(r['edades'][('Total_', str(e))] for e in range(14, 29))
              for r in registros if r['area'] == 'Total' and r['anio'] == 2024)
    pob = sum(r['total'] for r in registros if r['area'] == 'Total' and r['anio'] == 2024)
    print(f'\nVERIFICACIÓN — Bogotá 2024:')
    print(f'  Población total: {pob:,.0f}'.replace(',', '.'))
    print(f'  Jóvenes (14-28): {jov:,.0f}'.replace(',', '.'))
    print(f'\nFuente estándar lista. Ahora correr: python dim1/actualizar.py')


if __name__ == '__main__':
    main()
