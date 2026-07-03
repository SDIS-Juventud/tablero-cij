# crear_excel_datos.py — Dimensión 4: Salud integral y autocuidado
# Inicializa datos-dim4.xlsx desde los JSON existentes.
# Corre UNA SOLA VEZ para crear el archivo.
# A partir de ahí, la persona que actualiza solo abre el Excel y llena los datos nuevos.
#
# Uso: python dim4/crear_excel_datos.py
# Requiere: pip install openpyxl

import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR    = os.path.join(SCRIPT_DIR, 'data')
OUTPUT_PATH = os.path.join(os.path.dirname(SCRIPT_DIR),
                           'tablero-cij-compartido', 'dim4', 'datos-dim4.xlsx')

# ---- Colores ----
AZUL_OSC  = '1a4a7a'
AZUL      = '3a7fc1'
ROJO      = 'b5302a'
NARANJA   = 'e07830'
GRIS_CLARO = 'f4f5f7'
BLANCO    = 'FFFFFF'


def estilo_encabezado(ws, fila, columnas, color_fondo=AZUL_OSC):
    fill = PatternFill('solid', fgColor=color_fondo)
    font = Font(bold=True, color=BLANCO, size=10)
    for col in range(1, columnas + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def borde_fino():
    s = Side(style='thin', color='cccccc')
    return Border(left=s, right=s, top=s, bottom=s)


def zebra(ws, fila, n_cols):
    fill = PatternFill('solid', fgColor=GRIS_CLARO)
    if fila % 2 == 0:
        for col in range(1, n_cols + 1):
            ws.cell(row=fila, column=col).fill = fill


def meta_anio(ws, label, valor, color=AZUL):
    """Escribe 'label | valor' en A1:B1 como celda de referencia de año."""
    ws['A1'] = label
    ws['B1'] = valor
    ws['A1'].font = Font(bold=True, color=BLANCO, size=10)
    ws['B1'].font = Font(bold=True, color=BLANCO, size=10)
    ws['A1'].fill = PatternFill('solid', fgColor=color)
    ws['B1'].fill = PatternFill('solid', fgColor=color)
    ws['A1'].alignment = Alignment(horizontal='right')
    ws['B1'].alignment = Alignment(horizontal='left')


# ============================================================
# Hoja Fec_Bogota
# ============================================================

def llenar_fec_bogota(wb, datos_fec):
    ws = wb.create_sheet('Fec_Bogota')
    ws.sheet_properties.tabColor = ROJO

    headers = [
        'Año',
        'NV_10_14', 'Tasa_10_14',
        'NV_15_19', 'Tasa_15_19',
        'NV_20_24', 'Tasa_20_24',
        'NV_25_29', 'Tasa_25_29',
    ]
    ws.append(headers)
    estilo_encabezado(ws, 1, len(headers), ROJO)

    for r in datos_fec['por_anio']:
        ws.append([
            r['anio'],
            r.get('nv_10_14'), r.get('tasa_10_14'),
            r.get('nv_15_19'), r.get('tasa_15_19'),
            r.get('nv_20_24'), r.get('tasa_20_24'),
            r.get('nv_25_29'), r.get('tasa_25_29'),
        ])
        fila = ws.max_row
        zebra(ws, fila, len(headers))

    ws.column_dimensions['A'].width = 8
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = 'B2'
    print('  [OK] Fec_Bogota')


# ============================================================
# Hoja Fec_Localidad
# ============================================================

def llenar_fec_localidad(wb, datos_fec):
    ws = wb.create_sheet('Fec_Localidad')
    ws.sheet_properties.tabColor = NARANJA

    meta_anio(ws, 'Año de referencia:', datos_fec['ultimo_anio'], NARANJA)

    headers = [
        'Localidad',
        'NV_10_14', 'Tasa_10_14',
        'NV_15_19', 'Tasa_15_19',
        'NV_20_24', 'Tasa_20_24',
        'NV_25_29', 'Tasa_25_29',
    ]
    ws.append(headers)
    estilo_encabezado(ws, 2, len(headers), NARANJA)

    for r in datos_fec['por_localidad']:
        ws.append([
            r['localidad'],
            r.get('nv_10_14'), r.get('tasa_10_14'),
            r.get('nv_15_19'), r.get('tasa_15_19'),
            r.get('nv_20_24'), r.get('tasa_20_24'),
            r.get('nv_25_29'), r.get('tasa_25_29'),
        ])
        fila = ws.max_row
        zebra(ws, fila, len(headers))

    ws.column_dimensions['A'].width = 20
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = 'B3'
    print('  [OK] Fec_Localidad')


# ============================================================
# Hoja Mort_Bogota
# ============================================================

def llenar_mort_bogota(wb, datos_mort):
    ws = wb.create_sheet('Mort_Bogota')
    ws.sheet_properties.tabColor = AZUL_OSC

    headers = ['Año', 'Total', 'Hombres', 'Mujeres', 'Tasa_100k']
    ws.append(headers)
    estilo_encabezado(ws, 1, len(headers), AZUL_OSC)

    for r in datos_mort['por_anio']:
        ws.append([
            r['anio'],
            r.get('total'),
            r.get('hombres'),
            r.get('mujeres'),
            r.get('tasa_100k'),
        ])
        fila = ws.max_row
        zebra(ws, fila, len(headers))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    ws.freeze_panes = 'B2'
    print('  [OK] Mort_Bogota')


# ============================================================
# Hoja Mort_Causas
# ============================================================

def llenar_mort_causas(wb, datos_mort):
    ws = wb.create_sheet('Mort_Causas')
    ws.sheet_properties.tabColor = AZUL

    ultimo = datos_mort['por_anio'][-1]
    meta_anio(ws, 'Año de referencia:', ultimo['anio'], AZUL)

    headers = ['Causa (lista 6/67)', 'Casos']
    ws.append(headers)
    estilo_encabezado(ws, 2, len(headers), AZUL)

    for causa in ultimo.get('top_causas', []):
        ws.append([causa['causa'], causa['casos']])
        fila = ws.max_row
        zebra(ws, fila, len(headers))

    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 10

    ws.freeze_panes = 'A3'
    print('  [OK] Mort_Causas')


# ============================================================
# Hoja Mort_Localidad
# ============================================================

def llenar_mort_localidad(wb, datos_mort):
    ws = wb.create_sheet('Mort_Localidad')
    ws.sheet_properties.tabColor = AZUL

    ultimo = datos_mort['por_anio'][-1]
    meta_anio(ws, 'Año de referencia:', ultimo['anio'], AZUL)

    headers = ['Localidad', 'Casos', 'Tasa_100k']
    ws.append(headers)
    estilo_encabezado(ws, 2, len(headers), AZUL)

    for loc in ultimo.get('por_localidad', []):
        ws.append([loc['localidad'], loc['casos'], loc.get('tasa_100k')])
        fila = ws.max_row
        zebra(ws, fila, len(headers))

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12

    ws.freeze_panes = 'B3'
    print('  [OK] Mort_Localidad')


# ============================================================
# Hoja GUÍA
# ============================================================

def llenar_guia(wb, datos_fec, datos_mort):
    ws = wb.create_sheet('GUÍA')
    ws.sheet_properties.tabColor = '4CAF50'

    titulo_fill = PatternFill('solid', fgColor='1a4a7a')
    titulo_font = Font(bold=True, color='FFFFFF', size=12)
    seccion_fill = PatternFill('solid', fgColor='3a7fc1')
    seccion_font = Font(bold=True, color='FFFFFF', size=10)
    normal_font  = Font(size=10)

    guia = [
        ('GUÍA DE ACTUALIZACIÓN — Dimensión 4: Salud integral y autocuidado', True, 'titulo'),
        ('', False, 'normal'),
        ('Este Excel es la fuente de datos del tablero CIJ (sección 4.3 y 4.4).', False, 'normal'),
        ('Cada año, la persona responsable actualiza este archivo y corre los scripts Python.', False, 'normal'),
        ('Los scripts generan los archivos JSON que alimentan el tablero publicado en GitHub Pages.', False, 'normal'),
        ('', False, 'normal'),

        ('── CÓMO ACTUALIZAR CADA AÑO ──────────────────────────────────', False, 'seccion'),
        ('', False, 'normal'),

        ('PASO 1: Actualizar fecundidad (hoja Fec_Bogota)', False, 'negrita'),
        ('  Fuente: SaludData – Observatorio de Salud de Bogotá', False, 'normal'),
        ('  URL: https://saluddata.saludbogota.gov.co', False, 'normal'),
        ('  Ruta: Observatorio > Indicadores > Salud Sexual y Reproductiva', False, 'normal'),
        ('        > Tasa específica de fecundidad por área', False, 'normal'),
        ('  Filtros: Año = [nuevo año] | Localidad = Bogotá D.C.', False, 'normal'),
        ('  Qué anotar: NV y Tasa para grupos 10-14, 15-19, 20-24, 25-29', False, 'normal'),
        ('  Acción: Agregar UNA FILA al final de la hoja Fec_Bogota.', False, 'normal'),
        ('', False, 'normal'),

        ('PASO 2: Actualizar fecundidad por localidad (hoja Fec_Localidad)', False, 'negrita'),
        ('  Misma fuente y ruta que en el Paso 1.', False, 'normal'),
        ('  Filtros: Año = [nuevo año] | Todas las localidades', False, 'normal'),
        ('  Acción: REEMPLAZAR todos los datos de la hoja Fec_Localidad.', False, 'normal'),
        ('           Actualizar celda B1 (año de referencia) con el nuevo año.', False, 'normal'),
        ('', False, 'normal'),

        ('PASO 3: Actualizar mortalidad Bogotá (hoja Mort_Bogota)', False, 'negrita'),
        ('  Fuente: SaludData – Causas de mortalidad', False, 'normal'),
        ('  Ruta: Observatorio > Indicadores > Demografía > Causas de mortalidad', False, 'normal'),
        ('  Filtros: Año = [nuevo año] | Localidad = Bogotá | Edad = 15-19 + 20-24 + 25-29', False, 'normal'),
        ('  Qué anotar: Total muertes, Hombres, Mujeres, Tasa por 100.000 jóvenes', False, 'normal'),
        ('  Acción: Agregar UNA FILA al final de la hoja Mort_Bogota.', False, 'normal'),
        ('', False, 'normal'),

        ('PASO 4: Actualizar causas de mortalidad (hoja Mort_Causas)', False, 'negrita'),
        ('  Misma fuente que Paso 3. Filtrar por lista 6/67.', False, 'normal'),
        ('  Acción: REEMPLAZAR todos los datos. Actualizar celda B1 (año).', False, 'normal'),
        ('          Ingresar las 10 causas más frecuentes con su conteo.', False, 'normal'),
        ('', False, 'normal'),

        ('PASO 5: Actualizar mortalidad por localidad (hoja Mort_Localidad)', False, 'negrita'),
        ('  Misma fuente, filtrar por localidad.', False, 'normal'),
        ('  Acción: REEMPLAZAR todos los datos. Actualizar celda B1 (año).', False, 'normal'),
        ('', False, 'normal'),

        ('── CÓMO CORRER LOS SCRIPTS ────────────────────────────────────', False, 'seccion'),
        ('', False, 'normal'),
        ('Abrir terminal (PowerShell o cmd) en la carpeta raíz del proyecto:', False, 'normal'),
        ('  cd "G:\\Mi unidad\\CH_projects\\SDIS\\tablero-cij"', False, 'normal'),
        ('', False, 'normal'),
        ('Actualizar tablero de fecundidad:', False, 'normal'),
        ('  python dim4/actualizar_fecundidad.py', False, 'normal'),
        ('', False, 'normal'),
        ('Actualizar tablero de mortalidad:', False, 'normal'),
        ('  python dim4/actualizar_mortalidad.py', False, 'normal'),
        ('', False, 'normal'),

        ('── ESTRUCTURA DE COLUMNAS ─────────────────────────────────────', False, 'seccion'),
        ('', False, 'normal'),
        ('NV_XX_XX   = Nacidos vivos del grupo de edad', False, 'normal'),
        ('Tasa_XX_XX = Tasa específica (por cada 1.000 mujeres del grupo)', False, 'normal'),
        ('Tasa_100k  = Muertes por cada 100.000 jóvenes de 15-29 años', False, 'normal'),
        ('Años con "p" al final (ej: 2025p) = datos preliminares', False, 'normal'),
        ('', False, 'normal'),

        ('── FUENTES OFICIALES ───────────────────────────────────────────', False, 'seccion'),
        ('', False, 'normal'),
        (f'Fecundidad — último año: {datos_fec["ultimo_anio"]}', False, 'normal'),
        (f'Fuente: {datos_fec["fuente"]}', False, 'normal'),
        (f'Nota: {datos_fec["nota"]}', False, 'normal'),
        ('', False, 'normal'),
        (f'Mortalidad — último año: {datos_mort["por_anio"][-1]["anio"]}', False, 'normal'),
        (f'Nota: {datos_mort["nota"]}', False, 'normal'),
    ]

    for texto, _, tipo in guia:
        ws.append([texto])
        fila = ws.max_row
        cell = ws.cell(row=fila, column=1)
        if tipo == 'titulo':
            cell.fill = titulo_fill
            cell.font = titulo_font
        elif tipo == 'seccion':
            cell.fill = seccion_fill
            cell.font = seccion_font
        elif tipo == 'negrita':
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions['A'].width = 80
    print('  [OK] GUIA')


# ============================================================
# Main
# ============================================================

# ============================================================
# Datos 2025p de mortalidad (calculados desde el CSV del OSB)
# Se agregan en la inicialización porque el mortalidad.json solo llega a 2024.
# ============================================================

MORT_2025P = {
    'anio': '2025p', 'total': 1516, 'hombres': 1095, 'mujeres': 421, 'tasa_100k': 82.3,
    'top_causas': [
        {'causa': 'Agresiones (homicidios)',                                          'casos': 452},
        {'causa': 'Accidentes de transporte terrestre',                               'casos': 237},
        {'causa': 'Lesiones autoinfligidas intencionalmente (suicidios)',              'casos': 149},
        {'causa': 'Signos, síntomas y afecciones mal definidas',                      'casos': 73},
        {'causa': 'Enfermedades del sistema nervioso, excepto meningitis',             'casos': 68},
        {'causa': 'Tumores malignos de otras localizaciones y de las no especificadas','casos': 50},
        {'causa': 'Leucemia',                                                          'casos': 43},
        {'causa': 'Resto de las enfermedades',                                         'casos': 43},
        {'causa': 'Infecciones respiratorias agudas',                                  'casos': 37},
        {'causa': 'Eventos de intención no determinada',                               'casos': 37},
    ],
    'por_localidad': [
        {'localidad': 'Santa Fe',            'casos': 37,  'tasa_100k': 144.1},
        {'localidad': 'Los Mártires',        'casos': 25,  'tasa_100k': 126.5},
        {'localidad': 'Ciudad Bolívar',      'casos': 198, 'tasa_100k': 119.3},
        {'localidad': 'Teusaquillo',         'casos': 36,  'tasa_100k': 118.2},
        {'localidad': 'Sumapaz',             'casos': 1,   'tasa_100k': 108.0},
        {'localidad': 'San Cristóbal',       'casos': 103, 'tasa_100k': 105.8},
        {'localidad': 'La Candelaria',       'casos': 6,   'tasa_100k': 105.1},
        {'localidad': 'Barrios Unidos',      'casos': 33,  'tasa_100k': 103.4},
        {'localidad': 'Rafael Uribe Uribe',  'casos': 91,  'tasa_100k': 100.6},
        {'localidad': 'Antonio Nariño',      'casos': 17,  'tasa_100k': 90.8},
        {'localidad': 'Bosa',                'casos': 168, 'tasa_100k': 90.9},
        {'localidad': 'Chapinero',           'casos': 30,  'tasa_100k': 88.8},
        {'localidad': 'Puente Aranda',       'casos': 47,  'tasa_100k': 84.8},
        {'localidad': 'Usme',                'casos': 86,  'tasa_100k': 82.4},
        {'localidad': 'Tunjuelito',          'casos': 34,  'tasa_100k': 80.6},
        {'localidad': 'Kennedy',             'casos': 184, 'tasa_100k': 74.7},
        {'localidad': 'Fontibón',            'casos': 58,  'tasa_100k': 63.7},
        {'localidad': 'Engativá',            'casos': 115, 'tasa_100k': 62.6},
        {'localidad': 'Usaquén',             'casos': 71,  'tasa_100k': 60.6},
        {'localidad': 'Suba',                'casos': 176, 'tasa_100k': 59.4},
    ],
}


def main():
    print('=' * 60)
    print('Creando datos-dim4.xlsx')
    print('=' * 60)

    with open(os.path.join(DATA_DIR, 'fecundidad_adolescente.json'), encoding='utf-8') as f:
        datos_fec = json.load(f)
    with open(os.path.join(DATA_DIR, 'mortalidad.json'), encoding='utf-8') as f:
        datos_mort = json.load(f)

    # Agregar 2025p si no está ya en el JSON
    anios_json = [str(r['anio']) for r in datos_mort['por_anio']]
    if '2025p' not in anios_json:
        datos_mort['por_anio'].append(MORT_2025P)
        datos_mort['ultimo_anio'] = '2025p'
        print('  2025p mortalidad agregado desde datos calculados del CSV')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    llenar_fec_bogota(wb, datos_fec)
    llenar_fec_localidad(wb, datos_fec)
    llenar_mort_bogota(wb, datos_mort)
    llenar_mort_causas(wb, datos_mort)
    llenar_mort_localidad(wb, datos_mort)
    llenar_guia(wb, datos_fec, datos_mort)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)

    print(f'\nArchivo creado: {OUTPUT_PATH}')
    print('\nPróximos pasos:')
    print('  1. Abrir el Excel y verificar los datos')
    print('  2. Correr actualizar_fecundidad.py para generar JSON desde el Excel')
    print('  3. Correr actualizar_mortalidad.py para generar JSON desde el Excel')


if __name__ == '__main__':
    main()
