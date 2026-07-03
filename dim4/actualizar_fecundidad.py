# actualizar_fecundidad.py — Dimensión 4: Salud integral y autocuidado
# Lee datos-dim4.xlsx (hojas Fec_Bogota y Fec_Localidad) y genera fecundidad_adolescente.json.
#
# Flujo de actualización anual:
#   1. Abrir tablero-cij-compartido/dim4/datos-dim4.xlsx
#   2. Agregar fila nueva en Fec_Bogota (Bogotá histórico)
#   3. Reemplazar datos en Fec_Localidad y actualizar celda B1 (año)
#   4. Correr: python dim4/actualizar_fecundidad.py
#
# Requiere: pip install openpyxl

import os
import json
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(os.path.dirname(SCRIPT_DIR),
                            'tablero-cij-compartido', 'dim4', 'datos-dim4.xlsx')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'data', 'fecundidad_adolescente.json')

CORTE  = '11/06/2026'  # actualizar cada año con la fecha de descarga desde SaludData
FUENTE = ('SaludData – Observatorio de Salud de Bogotá. '
          'Actualización fuente: 11/06/2026. p: preliminar.')
NOTA   = ('Tasa específica de fecundidad = nacidos vivos de madres del grupo de edad '
          'por cada 1.000 mujeres del mismo grupo. '
          'Numerador: DANE–RUAF–ND (SDS), series finales 2014–2023, 2024 preliminar. '
          'Denominador: proyecciones DANE-FONDANE-SDP, CNPV 2018.')


def a_float(v):
    if v is None:
        return None
    return round(float(v), 2)


def a_int(v):
    if v is None:
        return 0
    return int(v)


def leer_fec_bogota(wb):
    ws = wb['Fec_Bogota']
    por_anio = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            break
        por_anio.append({
            'anio':       str(row[0]),
            'nv_10_14':   a_int(row[1]),
            'tasa_10_14': a_float(row[2]),
            'nv_15_19':   a_int(row[3]),
            'tasa_15_19': a_float(row[4]),
            'nv_20_24':   a_int(row[5]),
            'tasa_20_24': a_float(row[6]),
            'nv_25_29':   a_int(row[7]),
            'tasa_25_29': a_float(row[8]),
        })
    return por_anio


def leer_fec_localidad(wb):
    ws = wb['Fec_Localidad']
    ultimo_anio = str(ws['B1'].value or '').strip()
    por_localidad = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            break
        por_localidad.append({
            'localidad':  str(row[0]),
            'nv_10_14':   a_int(row[1]),
            'tasa_10_14': a_float(row[2]) or 0.0,
            'nv_15_19':   a_int(row[3]),
            'tasa_15_19': a_float(row[4]) or 0.0,
            'nv_20_24':   a_int(row[5]),
            'tasa_20_24': a_float(row[6]) or 0.0,
            'nv_25_29':   a_int(row[7]),
            'tasa_25_29': a_float(row[8]) or 0.0,
        })
    return por_localidad, ultimo_anio


def main():
    print('=' * 60)
    print('Fecundidad — actualización desde Excel')
    print('=' * 60)
    print(f'Excel: {EXCEL_PATH}')

    if not os.path.exists(EXCEL_PATH):
        print('ERROR: No se encontró datos-dim4.xlsx')
        print('  → Correr primero: python dim4/crear_excel_datos.py')
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    por_anio = leer_fec_bogota(wb)
    por_localidad, ultimo_anio = leer_fec_localidad(wb)
    wb.close()

    if not por_anio:
        print('ERROR: Hoja Fec_Bogota vacía o sin datos.')
        return
    if not ultimo_anio:
        ultimo_anio = por_anio[-1]['anio']

    resultado = {
        'por_anio':      por_anio,
        'por_localidad': por_localidad,
        'ultimo_anio':   ultimo_anio,
        'corte':         CORTE,
        'fuente':        FUENTE,
        'nota':          NOTA,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    ult = por_anio[-1]
    print(f'\nAños: {[r["anio"] for r in por_anio]}')
    print(f'Localidades: {len(por_localidad)}')
    print(f'Último año: {ult["anio"]}')
    print(f'  10-14: {ult["nv_10_14"]} NV | {ult["tasa_10_14"]} ‰')
    print(f'  15-19: {ult["nv_15_19"]} NV | {ult["tasa_15_19"]} ‰')
    print(f'  20-24: {ult["nv_20_24"]} NV | {ult["tasa_20_24"]} ‰')
    print(f'  25-29: {ult["nv_25_29"]} NV | {ult["tasa_25_29"]} ‰')
    print(f'\nArchivo: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
