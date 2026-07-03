# actualizar_mortalidad.py — Dimensión 4: Salud integral y autocuidado
# Lee datos-dim4.xlsx (hojas Mort_Bogota, Mort_Causas, Mort_Localidad) y genera mortalidad.json.
#
# Flujo de actualización anual:
#   1. Abrir tablero-cij-compartido/dim4/datos-dim4.xlsx
#   2. Agregar fila nueva en Mort_Bogota (total, hombres, mujeres, tasa)
#   3. Reemplazar datos en Mort_Causas y Mort_Localidad; actualizar B1 (año) en ambas hojas
#   4. Correr: python dim4/actualizar_mortalidad.py
#
# Requiere: pip install openpyxl

import os
import json
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(os.path.dirname(SCRIPT_DIR),
                            'tablero-cij-compartido', 'dim4', 'datos-dim4.xlsx')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'data', 'mortalidad.json')

CORTE = '11/06/2026'  # actualizar cada año con la fecha de descarga desde SaludData
NOTA = ('Defunciones de jóvenes 15-29 años residentes en Bogotá. '
        'Tasa por cada 100.000 jóvenes del rango 15-29. '
        'Causas agrupadas según lista 6/67. '
        'Fuente: DANE–RUAF–ND / Observatorio de Salud de Bogotá – SaludData. '
        'p: preliminar.')


def a_float(v):
    if v is None:
        return None
    return round(float(v), 1)


def a_int(v):
    if v is None:
        return 0
    return int(v)


def leer_mort_bogota(wb):
    ws = wb['Mort_Bogota']
    por_anio = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            break
        por_anio.append({
            'anio':     str(row[0]),
            'total':    a_int(row[1]),
            'hombres':  a_int(row[2]),
            'mujeres':  a_int(row[3]),
            'tasa_100k': a_float(row[4]),
        })
    return por_anio


def leer_mort_causas(wb):
    ws = wb['Mort_Causas']
    ultimo_anio = str(ws['B1'].value or '').strip()
    causas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            break
        causas.append({'causa': str(row[0]), 'casos': a_int(row[1])})
    return causas, ultimo_anio


def leer_mort_localidad(wb):
    ws = wb['Mort_Localidad']
    locs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            break
        locs.append({
            'localidad': str(row[0]),
            'casos':     a_int(row[1]),
            'tasa_100k': a_float(row[2]),
        })
    return locs


def main():
    print('=' * 60)
    print('Mortalidad — actualización desde Excel')
    print('=' * 60)
    print(f'Excel: {EXCEL_PATH}')

    if not os.path.exists(EXCEL_PATH):
        print('ERROR: No se encontró datos-dim4.xlsx')
        print('  → Correr primero: python dim4/crear_excel_datos.py')
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    por_anio   = leer_mort_bogota(wb)
    causas, ultimo_anio = leer_mort_causas(wb)
    locs       = leer_mort_localidad(wb)
    wb.close()

    if not por_anio:
        print('ERROR: Hoja Mort_Bogota vacía o sin datos.')
        return
    if not ultimo_anio:
        ultimo_anio = por_anio[-1]['anio']

    # Agregar causas y localidad al último año
    registros = por_anio.copy()
    registros[-1]['top_causas']   = causas
    registros[-1]['por_localidad'] = locs

    # Años anteriores no tienen causas ni localidad en el Excel → lista vacía
    for r in registros[:-1]:
        if 'top_causas' not in r:
            r['top_causas'] = []
        if 'por_localidad' not in r:
            r['por_localidad'] = []

    resultado = {
        'por_anio':   registros,
        'ultimo_anio': ultimo_anio,
        'corte':       CORTE,
        'nota':        NOTA,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    ult = por_anio[-1]
    print(f'\nAños: {[r["anio"] for r in por_anio]}')
    print(f'Último año: {ult["anio"]} — {ult["total"]:,} muertes | tasa {ult["tasa_100k"]}/100k')
    print(f'Causas registradas: {len(causas)}')
    print(f'Localidades: {len(locs)}')
    if causas:
        print(f'#1 causa: {causas[0]["causa"]} ({causas[0]["casos"]})')
    print(f'\nArchivo: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
