# actualizar_natalidad.py — Dimensión 4: Salud integral y autocuidado
# Script para procesar el CSV de natalidad del OSB y el Excel de
# proyecciones de población del DANE, y generar el JSON de tasas
# de natalidad en jóvenes.
#
# Uso:
#   1. Descargar datos de natalidad de SaludData / OSB
#   2. Guardar CSV en dim4/fuentes/natalidad/
#   3. El Excel de proyecciones ya debe estar en la misma carpeta
#   4. Correr: python dim4/actualizar_natalidad.py
#
# Dependencia: openpyxl (pip install openpyxl)

import os
import csv
import json
import openpyxl

# ============================================================
# Configuración
# ============================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUENTES_DIR = os.path.join(SCRIPT_DIR, 'fuentes', 'natalidad')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')

# Rango de edad de madres jóvenes (política de juventud)
EDAD_MIN = 14
EDAD_MAX = 28

# Grupos quinquenales para el filtro del tablero
GRUPOS_EDAD = [
    ('15 a 19 años', 15, 19),
    ('20 a 24 años', 20, 24),
    ('25 a 28 años', 25, 28),
]

# Solo datos desde 2018 (corte de la política)
ANIO_MIN = 2018


# ============================================================
# Funciones — Población (Excel DANE)
# ============================================================

def encontrar_excel(carpeta):
    """Busca el Excel de proyecciones de población."""
    for f in os.listdir(carpeta):
        if f.endswith('.xlsx') and 'proyecciones' in f.lower() and not f.startswith('~$'):
            return os.path.join(carpeta, f)
    return None


def cargar_poblacion(ruta_excel):
    """
    Lee el Excel de proyecciones DANE y extrae población de mujeres
    jóvenes por localidad y año.

    Retorna: dict[anio][localidad] = {grupo_edad: poblacion, 'total': N}
    También: dict[anio]['Bogotá'] con totales agregados
    """
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb['Localidades_edad_quinquenal']

    # Encontrar fila de encabezados
    headers = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        vals = [cell.value for cell in row]
        if 'COD_LOC' in vals:
            headers = vals
            header_row = i
            break

    if not headers:
        print('ERROR: No se encontró encabezado en el Excel')
        return {}

    # Índices de columnas de mujeres jóvenes
    # Mujeres_15-19, Mujeres_20-24, Mujeres_25-29
    col_indices = {}
    for idx, h in enumerate(headers):
        if h and isinstance(h, str):
            if h == 'Mujeres_15-19':
                col_indices['15-19'] = idx
            elif h == 'Mujeres_20-24':
                col_indices['20-24'] = idx
            elif h == 'Mujeres_25-29':
                col_indices['25-29'] = idx

    poblacion = {}  # anio -> localidad -> {grupo: N, total: N}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        area = row[2]  # AREA
        if area != 'Total':
            continue

        anio = row[3]  # AÑO
        if not isinstance(anio, (int, float)) or int(anio) < ANIO_MIN:
            continue
        anio = int(anio)

        localidad = str(row[1] or '').strip()  # NOM_LOC
        if not localidad:
            continue

        # Población de mujeres jóvenes por grupo quinquenal
        m15_19 = row[col_indices.get('15-19', 0)] or 0
        m20_24 = row[col_indices.get('20-24', 0)] or 0
        m25_29 = row[col_indices.get('25-29', 0)] or 0

        if anio not in poblacion:
            poblacion[anio] = {}

        poblacion[anio][localidad] = {
            '15-19': int(m15_19),
            '20-24': int(m20_24),
            '25-29': int(m25_29),
            'total': int(m15_19) + int(m20_24) + int(m25_29),
        }

    # Calcular total Bogotá por año
    for anio in poblacion:
        total_bog = {'15-19': 0, '20-24': 0, '25-29': 0, 'total': 0}
        for loc, vals in poblacion[anio].items():
            for k in total_bog:
                total_bog[k] += vals[k]
        poblacion[anio]['Bogotá'] = total_bog

    wb.close()
    return poblacion


# ============================================================
# Funciones — Natalidad (CSV OSB)
# ============================================================

def encontrar_csv(carpeta):
    """Busca el CSV de natalidad más reciente."""
    archivos = [f for f in os.listdir(carpeta)
                if f.endswith('.csv') and not f.startswith('~$')]
    if not archivos:
        return None
    archivos.sort()
    return os.path.join(carpeta, archivos[-1])


def asignar_grupo(edad):
    """Asigna un grupo quinquenal a la edad de la madre."""
    if 15 <= edad <= 19:
        return '15-19'
    elif 20 <= edad <= 24:
        return '20-24'
    elif 25 <= edad <= 28:
        return '25-29'  # Se usa 25-29 del DANE aunque el rango real es 25-28
    return None


def procesar_natalidad(ruta_csv):
    """
    Lee el CSV de natalidad y agrupa nacimientos de madres jóvenes
    por año, grupo de edad y localidad.
    """
    # nacidos[anio]['total'] = N
    # nacidos[anio][grupo] = N
    # nacidos_loc[anio][localidad] = N
    nacidos = {}
    nacidos_loc = {}

    with open(ruta_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            try:
                anio = int(row.get('ANO', '0'))
                edad = int(row.get('EDAD_MADRE', '0'))
                total = int(row.get('TOTAL_NACIDOS', '0'))
            except (ValueError, TypeError):
                continue

            if anio < ANIO_MIN or edad < EDAD_MIN or edad > EDAD_MAX:
                continue

            grupo = asignar_grupo(edad)
            if not grupo:
                continue

            # Ignorar filas de "00 - Bogotá" (es el total agregado)
            # y "Sin Dato" para evitar doble conteo
            localidad = row.get('LOCALIDAD_MADRE', '').strip()
            if localidad.startswith('00') or localidad == 'Sin Dato':
                continue

            # Por año y grupo (sumando solo localidades individuales)
            if anio not in nacidos:
                nacidos[anio] = {'total': 0, '15-19': 0, '20-24': 0, '25-29': 0}
            nacidos[anio]['total'] += total
            nacidos[anio][grupo] += total

            # Por localidad
            if anio not in nacidos_loc:
                nacidos_loc[anio] = {}
            loc_nombre = localidad.split(' - ', 1)[1] if ' - ' in localidad else localidad
            nacidos_loc[anio][loc_nombre] = nacidos_loc[anio].get(loc_nombre, 0) + total

    return nacidos, nacidos_loc


# ============================================================
# Generar JSON
# ============================================================

def generar_json(nacidos, nacidos_loc, poblacion):
    """Genera el JSON de natalidad con tasas por cada mil."""
    os.makedirs(DATA_DIR, exist_ok=True)

    anios = sorted(nacidos.keys())
    registros = []

    for anio in anios:
        n = nacidos[anio]
        pob = poblacion.get(anio, {}).get('Bogotá', {})

        registro = {
            'anio': anio,
            'nacidos_total': n['total'],
            'nacidos_15_19': n['15-19'],
            'nacidos_20_24': n['20-24'],
            'nacidos_25_28': n['25-29'],
        }

        # Tasas por cada mil mujeres jóvenes
        pob_total = pob.get('total', 0)
        if pob_total > 0:
            registro['tasa_total'] = round(n['total'] / pob_total * 1000, 2)
        else:
            registro['tasa_total'] = None

        for grupo, label in [('15-19', 'tasa_15_19'), ('20-24', 'tasa_20_24'), ('25-29', 'tasa_25_28')]:
            pob_g = pob.get(grupo, 0)
            if pob_g > 0:
                registro[label] = round(n[grupo] / pob_g * 1000, 2)
            else:
                registro[label] = None

        registros.append(registro)

    # Por localidad para el último año disponible
    ultimo_anio = anios[-1] if anios else None
    por_localidad = []
    if ultimo_anio and ultimo_anio in nacidos_loc:
        pob_anio = poblacion.get(ultimo_anio, {})
        for loc, nac in nacidos_loc[ultimo_anio].items():
            pob_loc = pob_anio.get(loc, {}).get('total', 0)
            tasa = round(nac / pob_loc * 1000, 2) if pob_loc > 0 else None
            por_localidad.append({
                'localidad': loc,
                'nacidos': nac,
                'poblacion_mujeres_jovenes': pob_loc,
                'tasa_por_mil': tasa,
            })
        por_localidad.sort(key=lambda x: x.get('tasa_por_mil') or 0, reverse=True)

    resultado = {
        'por_anio': registros,
        'por_localidad': por_localidad,
        'anio_localidad': ultimo_anio,
        'nota': 'Tasa = nacidos vivos de madres jóvenes / mujeres jóvenes × 1.000. Población del DANE (proyecciones CNPV 2018).',
    }

    ruta = os.path.join(DATA_DIR, 'natalidad.json')
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    print(f'  natalidad.json generado')
    return ruta


def main():
    print('=' * 60)
    print('Actualización — Dimensión 4: Natalidad')
    print('=' * 60)

    # 1. Cargar población
    excel = encontrar_excel(FUENTES_DIR)
    if not excel:
        print('ERROR: No se encontró Excel de proyecciones en', FUENTES_DIR)
        return
    print(f'  Proyecciones: {os.path.basename(excel)}')
    poblacion = cargar_poblacion(excel)
    print(f'  Años de población: {sorted(poblacion.keys())}')

    # 2. Cargar natalidad
    csv_file = encontrar_csv(FUENTES_DIR)
    if not csv_file:
        print('ERROR: No se encontró CSV de natalidad en', FUENTES_DIR)
        return
    print(f'  Natalidad: {os.path.basename(csv_file)}')
    nacidos, nacidos_loc = procesar_natalidad(csv_file)
    print(f'  Años de natalidad: {sorted(nacidos.keys())}')

    # 3. Generar JSON
    print('\nGenerando JSON...')
    ruta = generar_json(nacidos, nacidos_loc, poblacion)

    # Verificación
    with open(ruta, 'r', encoding='utf-8') as f:
        datos = json.load(f)

    print(f'\n{"=" * 60}')
    print('VERIFICACIÓN:')
    print(f'{"=" * 60}')
    for r in datos['por_anio']:
        tasa = f'{r["tasa_total"]:.1f}‰' if r['tasa_total'] else 'N/D'
        print(f'  {r["anio"]}: {r["nacidos_total"]:,} nacidos — tasa {tasa}'.replace(',', '.'))

    print(f'\nLocalidades ({datos["anio_localidad"]}): {len(datos["por_localidad"])}')
    if datos['por_localidad']:
        top = datos['por_localidad'][0]
        print(f'  Mayor tasa: {top["localidad"]} ({top["tasa_por_mil"]}‰)')

    print(f'\nArchivo: {ruta}')


if __name__ == '__main__':
    main()
