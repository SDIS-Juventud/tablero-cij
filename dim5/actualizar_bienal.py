# actualizar_bienal.py — Dimensión 5: Cultura, Recreación y Deporte
# Lee datos-dim5.xlsx (hoja Bienal_Culturas) y genera bienal_culturas.json.
#
# Flujo de actualización (cada dos años — encuesta bienal):
#   1. Abrir tablero-cij-compartido/dim5/datos-dim5.xlsx
#   2. Actualizar los valores en la hoja Bienal_Culturas
#   3. Correr: python dim5/actualizar_bienal.py
#
# Requiere: pip install openpyxl

import os
import json
import openpyxl

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(os.path.dirname(SCRIPT_DIR),
                            'tablero-cij-compartido', 'dim5', 'datos-dim5.xlsx')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'data', 'bienal_culturas.json')

FUENTE = 'Encuesta Bienal de Culturas 2025 — Secretaría de Cultura, Recreación y Deporte de Bogotá'
NOTA   = ('Grupo de edad 13-28 años como proxy de juventud (Ley 1622/2013 define 14-28 años). '
          'Encuesta bienal: actualización cada dos años. '
          'Fuente: Observatorio de Culturas — Secretaría de Cultura, Recreación y Deporte.')
CORTE  = '2025'  # actualizar con el año de la encuesta


def a_float(v):
    if v is None:
        return None
    return round(float(v), 1)


def main():
    print('=' * 60)
    print('Encuesta Bienal de Culturas — actualización desde Excel')
    print('=' * 60)
    print(f'Excel: {EXCEL_PATH}')

    if not os.path.exists(EXCEL_PATH):
        print('ERROR: No se encontró datos-dim5.xlsx')
        print('  → Crear el Excel con la hoja Bienal_Culturas primero.')
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if 'Bienal_Culturas' not in wb.sheetnames:
        print('ERROR: No se encontró la hoja "Bienal_Culturas" en el Excel.')
        return

    ws = wb['Bienal_Culturas']
    datos = {r[0]: r[1:] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}

    # Estructura esperada en la hoja:
    # Col A: indicador (ej: "practica_actual_jovenes")
    # Col B: valor jóvenes 13-28
    # Col C: valor adultos 29-59
    # Col D: valor mayores 60+
    def v(key, col):
        row = datos.get(key)
        if row is None:
            return None
        return a_float(row[col])

    resultado = {
        'fuente': FUENTE,
        'poblacion_representada': 6763265,
        'corte': CORTE,
        'nota': NOTA,
        'practica_cultural_actual': {
            'pregunta': '¿Practica actualmente alguna actividad cultural, artística o creativa?',
            'jovenes_13_28':  v('practica_actual', 0),
            'adultos_29_59':  v('practica_actual', 1),
            'mayores_60_mas': v('practica_actual', 2),
        },
        'satisfaccion_oferta_distrital': {
            'pregunta': '¿Cuál es su nivel de satisfacción con las actividades culturales y artísticas que organizan las entidades del Distrito en su localidad? (escala 1-4)',
            'jovenes_13_28': {
                'nada':          v('sat_nada', 0),
                'poco':          v('sat_poco', 0),
                'satisfecho':    v('sat_sat', 0),
                'muy_satisfecho': v('sat_muy', 0),
            },
            'adultos_29_59': {
                'nada':          v('sat_nada', 1),
                'poco':          v('sat_poco', 1),
                'satisfecho':    v('sat_sat', 1),
                'muy_satisfecho': v('sat_muy', 1),
            },
            'mayores_60_mas': {
                'nada':          v('sat_nada', 2),
                'poco':          v('sat_poco', 2),
                'satisfecho':    v('sat_sat', 2),
                'muy_satisfecho': v('sat_muy', 2),
            },
        },
        'participacion_actividades': {
            'pregunta': 'En los últimos 12 meses, ¿cuál de las siguientes actividades practicó o realizó?',
            'actividades': []
        }
    }

    # Actividades: filas que empiezan con "act_"
    for key, vals in datos.items():
        if str(key).startswith('act_'):
            nombre = vals[3] if len(vals) > 3 and vals[3] else key.replace('act_', '')
            resultado['participacion_actividades']['actividades'].append({
                'nombre':        nombre,
                'jovenes_13_28': a_float(vals[0]),
                'adultos_29_59': a_float(vals[1]),
                'mayores_60_mas': a_float(vals[2]),
            })

    wb.close()

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    print(f'\nPráctica cultural — jóvenes: {resultado["practica_cultural_actual"]["jovenes_13_28"]}%')
    j = resultado['satisfaccion_oferta_distrital']['jovenes_13_28']
    sat = (j['satisfecho'] or 0) + (j['muy_satisfecho'] or 0)
    print(f'Satisfechos con oferta distrital — jóvenes: {sat}%')
    print(f'Actividades registradas: {len(resultado["participacion_actividades"]["actividades"])}')
    print(f'\nArchivo: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
