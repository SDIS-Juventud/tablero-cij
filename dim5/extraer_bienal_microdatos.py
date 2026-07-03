# extraer_bienal_microdatos.py — Dimensión 5: Cultura, Recreación y Deporte
#
# Lee los microdatos crudos de la Encuesta Bienal de Culturas 2025 (dos archivos:
# Cultura y Deportes), filtra jóvenes de 13 a 28 años, calcula porcentajes
# ponderados por el factor de expansión, y deja el resultado en un Excel
# intermedio (datos-dim5.xlsx) que es lo que después lee actualizar_bienal.py.
#
# Por qué dos pasos (este script + actualizar_bienal.py) y no uno solo:
# si en la próxima Bienal (2027) la fuente cambia de formato, solo hay que
# reescribir este extractor. El Excel intermedio y actualizar_bienal.py
# quedan iguales.
#
# Uso:
#   python dim5/extraer_bienal_microdatos.py
#
# Requiere: pip install openpyxl

import os
import openpyxl
from openpyxl import Workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FUENTES_DIR = os.path.join(SCRIPT_DIR, 'fuentes', 'Encuesta Bienal de Culturas')
CULTURA_PATH = os.path.join(FUENTES_DIR, 'Cultura', 'm195_datos_consolidados.xlsx')
DEPORTES_PATH = os.path.join(FUENTES_DIR, 'Deportes', 'm182_datos.xlsx')
SALIDA_PATH = os.path.join(os.path.dirname(SCRIPT_DIR), 'tablero-cij-compartido', 'dim5', 'datos-dim5.xlsx')

EDAD_MIN = 13
EDAD_MAX = 28

# Columnas confirmadas manualmente contra el codebook y los datos crudos (ver Notes/)
COL_CULTURA = {
    'edad': 'VU',        # D3
    'factor': 'XC',
    'sat_cultura': 'AZ',   # P1.1 — satisfacción oferta cultural del Distrito (1-4)
    'sat_deporte': 'BB',   # P1.3 — satisfacción oferta deportiva/recreativa del Distrito (1-4)
    'practica_actual': 'BF',  # P2 — Sí/No (código 1=Sí, 2=No)
}
# P3.1 a P3.17 — práctica cultural (columnas BG a BW)
COLS_PRACTICA = {
    'BG': 'Cantar, crear o componer música',
    'BH': 'Interpretar instrumentos',
    'BI': 'Escribir (novela, poesía, cuento, ensayo, cómic...)',
    'BJ': 'Expresarse verbalmente (narración, podcasts, stand up...)',
    'BK': 'Esculpir, tejer, bordar, tallar, confeccionar',
    'BL': 'Pintar, dibujar, fotografiar',
    'BM': 'Diseñar instalaciones y escenografías',
    'BN': 'Bailar',
    'BO': 'Dirigir o actuar en espectáculos escénicos',
    'BP': 'Filmar, animar o editar contenidos audiovisuales',
    'BQ': 'Diseñar o crear contenidos publicitarios',
    'BR': 'Diseñar o fabricar juguetes',
    'BS': 'Diseñar o fabricar instrumentos musicales',
    'BT': 'Diseñar o fabricar joyería',
    'BU': 'Crear videojuegos, apps o software',
    'BV': 'Cocinar o hacer bebidas como tradición cultural',
    'BW': 'Otra actividad artística o creativa propia',
}
# P19_0 a P19_14 — asistencia cultural (columnas EH a EV)
COLS_ASISTENCIA = {
    'EH': 'Recorridos por el centro histórico',
    'EI': 'Visitas a monumentos',
    'EJ': 'Museos',
    'EK': 'Festivales y eventos culturales',
    'EL': 'Espacios reconocidos como patrimonio',
    'EM': 'Plazas fundacionales',
    'EN': 'Plazas de mercado',
    'EO': 'Obras de arte en el espacio público',
    'EP': 'Corredores de murales y arte urbano',
    'EQ': 'Iglesias o centros religiosos',
    'ER': 'Actividades culturales para primera infancia',
    'ES': 'Manifestaciones de pueblos y comunidades étnicas',
    'ET': 'Salas de exposiciones o galerías de arte',
    'EU': 'Distritos Creativos de Bogotá',
    'EV': 'Recorridos por lugares naturales o rurales',
}

COL_DEPORTES = {
    'edad': 'E',       # EDAD
    'factor': 'MX',
    'filter_': 'MY',   # 'Selected' / 'Not Selected'
    'practica_actual': 'BT',  # P22 — Sí/No
    'razon_no_practica': 'DT',  # P29 — solo aplica si practica_actual = No
    'donde_actividad': 'FQ',    # P33
    'etapa_cambio': 'GH',       # P40
}
# Nota: la pregunta P30 ("de las siguientes actividades físicas, ¿cuáles realiza
# actualmente?") se descartó por completo. Su categoría "otro deporte sin
# aspiraciones de alto rendimiento" deja fuera a quien practica deporte a nivel
# competitivo, lo que vuelve engañosa la pregunta como medida de actividad física.


# ============================================================
# Utilidades
# ============================================================

def leer_hoja_datos(ruta):
    """Abre el archivo crudo y devuelve (encabezados por letra, filas)."""
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb['datos']
    col_letters = [c.column_letter for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return col_letters, filas


def valor(fila, col_letters, letra):
    return fila[col_letters.index(letra)]


def pct_categoria(filas, col_letters, col_edad, col_factor, col_pregunta,
                   edad_min=EDAD_MIN, edad_max=EDAD_MAX, excluir_valores=None,
                   filtro_extra=None):
    """
    Calcula el % ponderado de cada categoría de una variable categórica,
    sobre la base de jóvenes en el rango de edad (y cualquier filtro extra,
    ej. filter_ != 'Not Selected', o practica_actual == 'No').

    Devuelve una lista de (categoria, pct) ordenada de mayor a menor,
    y el total ponderado usado como base.
    """
    excluir_valores = excluir_valores or set()
    pesos_por_categoria = {}
    total = 0.0
    for fila in filas:
        edad = valor(fila, col_letters, col_edad)
        if edad is None:
            continue
        edad = int(edad)
        if edad < edad_min or edad > edad_max:
            continue
        if filtro_extra and not filtro_extra(fila, col_letters):
            continue
        cat = valor(fila, col_letters, col_pregunta)
        if cat is None or cat in excluir_valores:
            continue
        factor = valor(fila, col_letters, col_factor) or 0
        pesos_por_categoria[cat] = pesos_por_categoria.get(cat, 0.0) + factor
        total += factor

    if total == 0:
        return [], 0.0
    resultado = [(cat, round(peso / total * 100, 1)) for cat, peso in pesos_por_categoria.items()]
    resultado.sort(key=lambda x: -x[1])
    return resultado, total


def pct_binaria(filas, col_letters, col_edad, col_factor, col_pregunta,
                 valores_si, edad_min=EDAD_MIN, edad_max=EDAD_MAX, filtro_extra=None):
    """% ponderado de 'Sí' (o equivalente) sobre el total de jóvenes en el rango de edad."""
    peso_si = 0.0
    total = 0.0
    for fila in filas:
        edad = valor(fila, col_letters, col_edad)
        if edad is None:
            continue
        edad = int(edad)
        if edad < edad_min or edad > edad_max:
            continue
        if filtro_extra and not filtro_extra(fila, col_letters):
            continue
        factor = valor(fila, col_letters, col_factor) or 0
        total += factor
        cat = valor(fila, col_letters, col_pregunta)
        if cat in valores_si:
            peso_si += factor
    if total == 0:
        return 0.0, 0
    return round(peso_si / total * 100, 1), total


def pct_practica_p3(filas, col_letters, col_edad, col_factor, col_pregunta,
                     edad_min=EDAD_MIN, edad_max=EDAD_MAX):
    """
    Para las variables P3.x de Cultura: código 1 o 2 = practica (con alguna
    frecuencia), código 3 = no realiza, vacío = no aplica (se cuenta como
    no practica, igual que en el tablero anterior).
    """
    peso_practica = 0.0
    total = 0.0
    for fila in filas:
        edad = valor(fila, col_letters, col_edad)
        if edad is None:
            continue
        edad = int(edad)
        if edad < edad_min or edad > edad_max:
            continue
        factor = valor(fila, col_letters, col_factor) or 0
        total += factor
        cat = valor(fila, col_letters, col_pregunta)
        if cat in (1, 2):
            peso_practica += factor
    if total == 0:
        return 0.0
    return round(peso_practica / total * 100, 1)


# ============================================================
# Cultura
# ============================================================

def procesar_cultura():
    print(f'Leyendo {os.path.basename(CULTURA_PATH)}...')
    col_letters, filas = leer_hoja_datos(CULTURA_PATH)

    ce, cf = COL_CULTURA['edad'], COL_CULTURA['factor']

    practica_actual, n_practica = pct_binaria(
        filas, col_letters, ce, cf, COL_CULTURA['practica_actual'], valores_si={1})

    sat_cultura, _ = pct_categoria(filas, col_letters, ce, cf, COL_CULTURA['sat_cultura'])
    sat_deporte, _ = pct_categoria(filas, col_letters, ce, cf, COL_CULTURA['sat_deporte'])

    etiquetas_sat = {1: 'Nada satisfecho/a', 2: 'Poco satisfecho/a',
                     3: 'Satisfecho/a', 4: 'Muy satisfecho/a'}
    sat_cultura = {etiquetas_sat[c]: p for c, p in sat_cultura}
    sat_deporte = {etiquetas_sat[c]: p for c, p in sat_deporte}

    practica = []
    for col, nombre in COLS_PRACTICA.items():
        pct = pct_practica_p3(filas, col_letters, ce, cf, col)
        practica.append((nombre, pct))
    practica.sort(key=lambda x: -x[1])

    asistencia = []
    for col, nombre in COLS_ASISTENCIA.items():
        pct, _ = pct_binaria(filas, col_letters, ce, cf, col, valores_si={1})
        asistencia.append((nombre, pct))
    asistencia.sort(key=lambda x: -x[1])

    print(f'  Jóvenes 13-28 (Cultura): {n_practica:,.0f} personas ponderadas'.replace(',', '.'))
    print(f'  Práctica cultural actual: {practica_actual}%')

    return {
        'practica_actual': practica_actual,
        'n_jovenes': n_practica,
        'satisfaccion_cultura': sat_cultura,
        'satisfaccion_deporte': sat_deporte,
        'practica': practica,
        'asistencia': asistencia,
    }


# ============================================================
# Deportes
# ============================================================

def filtro_seleccionado(fila, col_letters):
    return valor(fila, col_letters, COL_DEPORTES['filter_']) != 'Not Selected'


def procesar_deportes():
    print(f'Leyendo {os.path.basename(DEPORTES_PATH)}...')
    col_letters, filas_todas = leer_hoja_datos(DEPORTES_PATH)

    n_excluidas = sum(
        1 for f in filas_todas
        if valor(f, col_letters, COL_DEPORTES['filter_']) == 'Not Selected'
    )
    print(f'  Filas excluidas por filter_ = "Not Selected": {n_excluidas}')

    ce, cf = COL_DEPORTES['edad'], COL_DEPORTES['factor']

    practica_actual, n_practica = pct_binaria(
        filas_todas, col_letters, ce, cf, COL_DEPORTES['practica_actual'],
        valores_si={'Sí'}, filtro_extra=filtro_seleccionado)

    def filtro_no_practica(fila, col_letters):
        return (filtro_seleccionado(fila, col_letters)
                and valor(fila, col_letters, COL_DEPORTES['practica_actual']) == 'No')

    razon_no_practica, _ = pct_categoria(
        filas_todas, col_letters, ce, cf, COL_DEPORTES['razon_no_practica'],
        filtro_extra=filtro_no_practica)

    # Nota: "Necesito tomar clases" y "Tengo varias obligaciones..." también
    # aparecen como respuesta de esta pregunta (P33), con el mismo texto que
    # dos opciones de P29. Se verificó que la mayoría de quienes respondieron
    # esto sí reportan practicar alguna actividad física en P30, así que no se
    # excluyen: no hay evidencia suficiente de que sea un error de los datos.
    donde_actividad, _ = pct_categoria(
        filas_todas, col_letters, ce, cf, COL_DEPORTES['donde_actividad'],
        filtro_extra=filtro_seleccionado)

    etapa_cambio, _ = pct_categoria(
        filas_todas, col_letters, ce, cf, COL_DEPORTES['etapa_cambio'],
        filtro_extra=filtro_seleccionado)

    print(f'  Jóvenes 13-28 (Deportes): {n_practica:,.0f} personas ponderadas'.replace(',', '.'))
    print(f'  Práctica deportiva actual: {practica_actual}%')

    return {
        'practica_actual': practica_actual,
        'n_jovenes': n_practica,
        'n_excluidas_not_selected': n_excluidas,
        'razon_no_practica': razon_no_practica,
        'donde_actividad': donde_actividad,
        'etapa_cambio': etapa_cambio,
    }


# ============================================================
# Escritura del Excel intermedio
# ============================================================

def escribir_hoja_lista(wb, nombre, encabezados, filas):
    ws = wb.create_sheet(nombre)
    ws.append(encabezados)
    for fila in filas:
        ws.append(fila)


def generar_excel(cultura, deportes):
    os.makedirs(os.path.dirname(SALIDA_PATH), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    ws_info = wb.create_sheet('Info')
    ws_info.append(['Campo', 'Valor'])
    filas_info = [
        ('Fuente Cultura', 'Encuesta Bienal de Culturas 2025 — Secretaría de Cultura, Recreación y Deporte'),
        ('Fuente Deportes', 'Encuesta de Prácticas Deportivas y Calidad de Vida 2024 (módulo de la Bienal)'),
        ('Corte edad jóvenes', f'{EDAD_MIN} a {EDAD_MAX} años'),
        ('Jóvenes ponderados (Cultura)', round(cultura['n_jovenes'])),
        ('Jóvenes ponderados (Deportes)', round(deportes['n_jovenes'])),
        ('Filas excluidas Deportes (filter_ = Not Selected)', deportes['n_excluidas_not_selected']),
        ('Categorías de participación cultural descontinuadas',
         'Ferias, Conferencias, Tertulias literarias y Carnaval ya no se preguntan en la encuesta 2025'),
        ('Oferta privada (P1.2/P1.4)', 'No se incluye, solo se usa oferta del Distrito'),
    ]
    for f in filas_info:
        ws_info.append(f)

    ws_kpi = wb.create_sheet('KPIs')
    ws_kpi.append(['indicador', 'valor'])
    ws_kpi.append(['practica_cultural_actual', cultura['practica_actual']])
    ws_kpi.append(['practica_deportiva_actual', deportes['practica_actual']])

    ws_sat = wb.create_sheet('Satisfaccion')
    ws_sat.append(['nivel', 'cultura_distrito', 'deporte_distrito'])
    for nivel in ['Nada satisfecho/a', 'Poco satisfecho/a', 'Satisfecho/a', 'Muy satisfecho/a']:
        ws_sat.append([nivel, cultura['satisfaccion_cultura'].get(nivel, 0),
                        cultura['satisfaccion_deporte'].get(nivel, 0)])

    escribir_hoja_lista(wb, 'Practica_Cultural', ['actividad', 'pct'], cultura['practica'])
    escribir_hoja_lista(wb, 'Asistencia_Cultural', ['actividad', 'pct'], cultura['asistencia'])
    escribir_hoja_lista(wb, 'Razon_No_Deporte', ['razon', 'pct'], deportes['razon_no_practica'])
    escribir_hoja_lista(wb, 'Donde_Actividad_Fisica', ['lugar', 'pct'], deportes['donde_actividad'])
    escribir_hoja_lista(wb, 'Etapas_Cambio', ['etapa', 'pct'], deportes['etapa_cambio'])

    wb.save(SALIDA_PATH)
    print(f'\nExcel intermedio generado: {SALIDA_PATH}')


def main():
    print('=' * 60)
    print('Extracción de microdatos — Dimensión 5 (Bienal de Culturas)')
    print('=' * 60)
    cultura = procesar_cultura()
    print()
    deportes = procesar_deportes()
    print()
    generar_excel(cultura, deportes)


if __name__ == '__main__':
    main()
